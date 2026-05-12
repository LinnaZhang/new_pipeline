import openpyxl
from .utils import column_letter_to_number

def apply_formula(ws, config):
    """应用自定义公式到指定列 (基于内存里的 ws 对象)"""
    target_cols = config.get('target_cols', [config.get('target_col')])
    formula_type = config['formula_type']
    start_row = config['start_row']
    format_str = config.get('format', '0.00%')
    
    # 自动检测结束行
    end_row = config.get('end_row')
    if end_row is None:
        if formula_type in ['yoy', 'mom', 'custom2']:
            # 对于同比/环比，检查A列或源数据列的最大有效行
            # 我们假设A列是日期列，或者我们可以遍历一下
            for row in range(ws.max_row, start_row - 1, -1):
                if ws.cell(row=row, column=1).value is not None:
                    end_row = row
                    break
        if end_row is None:
            end_row = ws.max_row
            
    params = config.get('params', {})
    
    for target_col in target_cols:
        target_col_num = column_letter_to_number(target_col)
        
        # 对于多个目标列，我们需要根据偏移量或者直接在params中指定source_col
        # 在这里我们为了兼容，如果是批量传入 target_cols，我们可以要求 params 里面定义如何找到 source_col
        # 例如 source_offset = -1 表示 source_col 是 target_col 左边一列
        if 'source_offset' in params:
            source_col_num = target_col_num + params['source_offset']
            source_col = openpyxl.utils.get_column_letter(source_col_num)
        else:
            source_col = params.get('source_col')
            
        for row in range(start_row, end_row + 1):
            formula = ""
            try:
                date_value = str(ws.cell(row=row, column=1).value) if ws.cell(row=row, column=1).value else ""
                
                if formula_type == 'yoy':
                    periods = params.get('periods', 12)
                    
                    # 针对年度数据的特殊处理 (来自航空逻辑)
                    if params.get('year_condition') and "年" in date_value:
                        if params.get('skip_2018') and "2018年" in date_value:
                            continue
                        prev_row = row - params.get('year_periods', 2)
                        if prev_row >= 1:
                            formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                    else:
                        prev_row = row - periods
                        if prev_row >= 1:
                            formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                            
                elif formula_type == 'custom2':
                    # 同比19 的特殊公式 (从老代码迁移过来的简化版)
                    # 航空月报的特殊逻辑: 使用 "2019年" 或 2019 年对应月份的数据
                    pass  # 这里可以抽象到插件，或者如果在 YAML 里直接用插件 action
                    
            except Exception:
                pass

            if formula:
                cell = ws.cell(row=row, column=target_col_num)
                cell.value = formula
                cell.number_format = format_str

    print(f"已将通用公式 {formula_type} 应用到 {target_cols}")
