import pandas as pd
import numpy as np
from typing import Dict, List, Union, Optional, Tuple
import logging
from datetime import datetime, timedelta
from core_engine.data_reader import DataReader
import openpyxl.styles
from openpyxl.styles import Font, PatternFill, Alignment


# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def column_letter_to_number(column_letter):
    """将Excel列字母转换为数字索引（A->1, B->2, AA->27等）"""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def column_number_to_letter(column_number):
    """将数字索引转换为Excel列字母（1->A, 2->B, 27->AA等）"""
    result = ""
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

class DataProcessor:
    """
    负责处理数据和执行衍生计算的类
    """
    
    def __init__(self, data_reader: DataReader):
        """
        初始化DataProcessor
        
        Args:
            data_reader: DataReader实例，用于读取原始数据
        """
        self.data_reader = data_reader
    
    def filter_by_date_range(self, indicator_data: Dict, start_date: str = None, end_date: str = None) -> Dict:
        """
        筛选指定时间范围内的指标数据
        
        Args:
            indicator_data: 指标数据字典，包含metadata和data
            start_date: 开始日期，格式为'YYYY-MM-DD'，如果为None则不限制开始日期
            end_date: 结束日期，格式为'YYYY-MM-DD'，如果为None则不限制结束日期
            
        Returns:
            筛选后的指标数据字典
        """
        try:
            # 复制原始数据，避免修改原始数据
            result = {
                "metadata": indicator_data["metadata"].copy(),
                "data": indicator_data["data"].copy()
            }
            
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(result["data"]["日期"]):
                result["data"]["日期"] = pd.to_datetime(result["data"]["日期"], errors='coerce')
            
            # 筛选日期范围
            mask = pd.Series(True, index=result["data"].index)
            
            if start_date:
                start_date = pd.to_datetime(start_date)
                mask = mask & (result["data"]["日期"] >= start_date)
                
            if end_date:
                end_date = pd.to_datetime(end_date)
                mask = mask & (result["data"]["日期"] <= end_date)
            
            result["data"] = result["data"][mask].reset_index(drop=True)
            
            logger.info(f"已筛选时间范围: {start_date} 至 {end_date}, 剩余数据行数: {len(result['data'])}")
            
            return result
        
        except Exception as e:
            logger.error(f"筛选时间范围时出错: {str(e)}")
            raise
    
    def sort_by_date(self, indicator_data: Dict, ascending: bool = True) -> Dict:
        """
        按日期对指标数据进行排序
        
        Args:
            indicator_data: 指标数据字典，包含metadata和data
            ascending: 是否升序排列，True为升序，False为降序
            
        Returns:
            排序后的指标数据字典
        """
        try:
            # 复制原始数据，避免修改原始数据
            result = {
                "metadata": indicator_data["metadata"].copy(),
                "data": indicator_data["data"].copy()
            }
            
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(result["data"]["日期"]):
                result["data"]["日期"] = pd.to_datetime(result["data"]["日期"], errors='coerce')
            
            # 按日期排序
            result["data"] = result["data"].sort_values(by="日期", ascending=ascending).reset_index(drop=True)
            
            order_type = "升序" if ascending else "降序"
            logger.info(f"已按日期{order_type}排列数据")
            
            return result
        
        except Exception as e:
            logger.error(f"按日期排序时出错: {str(e)}")
            raise
    
    def create_month_year_pivot(self, indicator_data: Dict, value_column: str = None) -> pd.DataFrame:
        """
        创建以月份为横坐标、年份为纵坐标的交叉表
        
        Args:
            indicator_data: 指标数据字典，包含metadata和data
            value_column: 要展示的值列名，如果为None则使用第二列（通常是指标值）
            
        Returns:
            交叉表DataFrame
        """
        try:
            # 获取数据
            data = indicator_data["data"].copy()
            
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(data["日期"]):
                data["日期"] = pd.to_datetime(data["日期"], errors='coerce')
            
            # 提取年份和月份，但只保留A列内容
            if 'A' not in data.columns:
                data["年份"] = data["日期"].dt.year
                data["月份"] = data["日期"].dt.month
            # 清空除A列外的其他列
            for col in data.columns:
                if col != 'A':
                    data[col] = None
            
            # 确定值列
            if value_column is None:
                # 使用第二列作为值列
                value_column = data.columns[1]
            
            # 创建交叉表
            pivot_table = pd.pivot_table(
                data,
                values=value_column,
                index="年份",
                columns="月份",
                aggfunc="mean"  # 如果同一年月有多个值，取平均值
            )
            
            # 重命名列，使用月份名称
            month_names = {
                1: "一月", 2: "二月", 3: "三月", 4: "四月",
                5: "五月", 6: "六月", 7: "七月", 8: "八月",
                9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
            }
            pivot_table = pivot_table.rename(columns=month_names)
            
            logger.info(f"已创建月份-年份交叉表，形状: {pivot_table.shape}")
            
            return pivot_table
        
        except Exception as e:
            logger.error(f"创建月份-年份交叉表时出错: {str(e)}")
            raise
    
    def calculate_year_over_year_change(self, indicator_data: Dict) -> Dict:
        """
        计算同比变化率
        
        Args:
            indicator_data: 指标数据字典，包含metadata和data
            
        Returns:
            包含原始数据和同比变化率的字典
        """
        try:
            # 复制原始数据，避免修改原始数据
            result = {
                "metadata": indicator_data["metadata"].copy(),
                "data": indicator_data["data"].copy()
            }
            
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(result["data"]["日期"]):
                result["data"]["日期"] = pd.to_datetime(result["data"]["日期"], errors='coerce')
            
            # 保持A列内容，清空其他列
            for col in result["data"].columns:
                if col != 'A':
                    result["data"][col] = None
            
            # 获取值列名（仅用于处理A列）
            if 'A' in result["data"].columns:
                value_column = 'A'
            else:
                value_column = result["data"].columns[1]
            
            # 仅对A列进行排序和处理
            if 'A' in result["data"].columns:
                result["data"] = result["data"].sort_values(by='A').reset_index(drop=True)
            for i, row in result["data"].iterrows():
                # 查找去年同期数据
                last_year = row["年份"] - 1
                same_month = row["月份"]
                
                last_year_data = result["data"][
                    (result["data"]["年份"] == last_year) & 
                    (result["data"]["月份"] == same_month)
                ]
                
                # 检查去年同期数据是否存在
                if last_year_data.empty:
                    # 如果去年同期数据不存在，设置同比变化率为空值
                    result["data"].at[i, "同比变化率"] = None
                else:
                    current_value = row[value_column]
                    last_year_value = last_year_data.iloc[0][value_column]
                    
                    # 只有当两个条件都满足时才计算同比变化率：
                    # 1. 当前值不为空
                    # 2. 去年同期值不为0
                    if pd.notna(current_value) and last_year_value != 0:
                        yoy_change = (current_value - last_year_value) / last_year_value
                        result["data"].at[i, "同比变化率"] = yoy_change
                    else:
                        # 如果任一条件不满足，设置为空值
                        result["data"].at[i, "同比变化率"] = None
            
            # 删除辅助列
            result["data"] = result["data"].drop(columns=["年份", "月份"])
            
            logger.info(f"已计算同比变化率")
            
            return result
        
        except Exception as e:
            logger.error(f"计算同比变化率时出错: {str(e)}")
            raise
    
    def add_quarterly_stats(self, ws, start_row: int, data_df: pd.DataFrame, value_cols: List[str]) -> Tuple[int, List[Dict]]:
        """
        在Excel工作表中添加季度统计数据，并计算季度同比变化率
        
        Args:
            ws: openpyxl的worksheet对象
            start_row: 数据开始的行号
            data_df: 包含日期列和数据列的DataFrame，日期列名为'日期'
            value_cols: 需要进行季度统计的列名列表
            
        Returns:
            Tuple[int, List[Dict]]: 返回季度标题行号和季度统计结果列表
        """
        try:
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(data_df['日期']):
                data_df['日期'] = pd.to_datetime(data_df['日期'])
            
            # 计算最后一行数据的位置和季度标题行的位置
            last_data_row = start_row + len(data_df) - 1
            quarter_title_row = last_data_row + 2
            
            
            # 准备按季度分组的数据
            quarterly_groups = {}
            current_quarter_months = set()
            current_quarter_data = {}
            
            # 遍历数据处理每个月
            for _, row in data_df.iterrows():
                date = row['日期']
                year = date.year
                month = date.month
                quarter = (month - 1) // 3 + 1  # 计算季度 (1-3: Q1, 4-6: Q2, 7-9: Q3, 10-12: Q4)
                quarter_key = (year, quarter)
                
                # 初始化当前季度数据
                if quarter_key not in quarterly_groups:
                    quarterly_groups[quarter_key] = {
                        'quarter_label': f"{str(year)[-2:]}Q{quarter}",
                        'data': {col: 0 for col in value_cols},
                        'months': set(),
                        'month_values': {col: [] for col in value_cols}  # 用于计算平均值的数据存储
                    }
                
                # 处理当月数据
                current_group = quarterly_groups[quarter_key]
                for col in value_cols:
                    if pd.notna(row[col]):  # 只处理非空值
                        if col in ['AL', 'AO', 'AR', 'AU']:
                            # 对于AL、AO、AR、AU列，存储每月的值用于计算平均值
                            current_group['month_values'][col].append(row[col])
                        else:
                            # 对于其他列，继续使用累加
                            current_group['data'][col] += row[col]
                current_group['months'].add(month)
            
            # 整理季度数据
            quarter_data = []
            prev_year = None
            
            # 按年份和季度排序
            sorted_quarters = sorted(quarterly_groups.keys())
            
            for year, quarter in sorted_quarters:
                current_group = quarterly_groups[(year, quarter)]
                
                # 只处理完整的季度（有3个月的数据）
                if len(current_group['months']) == 3:
                    # 如果年份变化，添加空行
                    if prev_year is not None and year != prev_year:
                        quarter_data.append({
                            'quarter_label': '',
                            'data': {col: None for col in value_cols},
                            'year': None,
                            'quarter': None
                        })
                    
                    # 计算同比变化率
                    last_year_key = (year - 1, quarter)
                    current_data = current_group['data'].copy()
                    
                    # 对于AL、AO、AR、AU列，计算平均值
                    for col in ['AL', 'AO', 'AR', 'AU']:
                        if col in current_group['month_values'] and current_group['month_values'][col]:
                            current_data[col] = sum(current_group['month_values'][col]) / len(current_group['month_values'][col])
                        else:
                            current_data[col] = 0
                    
                    # 如果有去年同期的数据，计算同比变化率
                    if last_year_key in quarterly_groups and len(quarterly_groups[last_year_key]['months']) == 3:
                        last_year_data = quarterly_groups[last_year_key]['data'].copy()
                        
                        # 对于去年同期的AL、AO、AR、AU列，也需要计算平均值
                        last_year_group = quarterly_groups[last_year_key]
                        for col in ['AL', 'AO', 'AR', 'AU']:
                            if col in last_year_group['month_values'] and last_year_group['month_values'][col]:
                                last_year_data[col] = sum(last_year_group['month_values'][col]) / len(last_year_group['month_values'][col])
                            else:
                                last_year_data[col] = 0
                        
                        for col in value_cols:
                            # 计算原始数据列的同比变化率（仅针对原始数据列，不处理已有的同比列）
                            if col in ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W', 'Z', 'AC', 'AF', 'AI', 'AL', 'AO', 'AR', 'AU', 'AX']:
                                col_num = column_letter_to_number(col)
                                yoy_col = column_number_to_letter(col_num + 1)  # 获取下一列作为同比列
                                if last_year_data[col] != 0 and last_year_data[col] is not None:
                                    # 季度同比上年率计算逻辑：该季度同比上年率=（上一年季度指标/该年季度指标）-1
                                    current_data[yoy_col] = (current_data[col] / last_year_data[col]) - 1
                                else:
                                    current_data[yoy_col] = None
                    
                    quarter_data.append({
                        'quarter_label': current_group['quarter_label'],
                        'data': current_data,
                        'year': year,
                        'quarter': quarter
                    })
                    
                prev_year = year
            
            return quarter_title_row, quarter_data
            
        except Exception as e:
            logger.error(f"添加季度统计时出错: {str(e)}")
            raise

    def add_ytd_stats(self, ws, last_quarter_row: int, data_df: pd.DataFrame, value_cols: List[str], start_row: int = 2) -> int:
        """
        在Excel工作表中添加年度累计（YTD）统计数据
        
        Args:
            ws: openpyxl的worksheet对象
            last_quarter_row: 最后一个季度数据的行号
            data_df: 包含日期列和数据列的DataFrame，日期列名为'日期'
            value_cols: 需要进行统计的列名列表
            start_row: 数据开始的行号，默认为2（假设第1行是标题）
            
        Returns:
            int: 返回YTD统计数据的最后一行行号
        """
        try:
            # 确保日期列是datetime类型
            if not pd.api.types.is_datetime64_any_dtype(data_df['日期']):
                data_df['日期'] = pd.to_datetime(data_df['日期'])
            
            # YTD标题行位置
            ytd_title_row = last_quarter_row + 2
            # 月份数设置行位置
            month_count_row = ytd_title_row + 1
            # 数据起始行位置
            ytd_data_start_row = month_count_row + 1
            
            # 写入YTD标题并记住YTD所在行
            ytd_formula = f'="YTD"'
            ws[f'A{ytd_title_row}'] = ytd_formula
            ytd_row = ytd_title_row  # 保存YTD所在行号
            ws[f'A{ytd_row}'].font = Font(bold=True)

            # 清空指定行（除A列外）的内容，并且清空A列为空值的行的所有内容
            for row in [ytd_title_row-1, ytd_title_row, ytd_title_row+1, ytd_title_row+2]:
                # 检查A列是否为空值
                if ws[f'A{row}'].value is None or ws[f'A{row}'].value == "":
                    # 如果A列为空，清空整行
                    max_column = len(value_cols) + 1
                    for col_num in range(1, max_column + 1):  # 从1开始，包含A列
                        col_letter = column_number_to_letter(col_num)
                        ws[f'{col_letter}{row}'] = None
                else:
                    # 如果A列不为空，只清空除A列外的其他列
                    max_column = len(value_cols) + 1
                    for col_num in range(2, max_column + 1):  # 从2开始，不包含A列
                        col_letter = column_number_to_letter(col_num)
                        ws[f'{col_letter}{row}'] = None
            
            # 动态检测最新数据月份并设置month_count_formula
            latest_date = None
            latest_month = 6  # 默认值
            
            # 查找最新的日期数据
            for _, row in data_df.iterrows():
                date = row['日期']
                if latest_date is None or date > latest_date:
                    latest_date = date
            
            if latest_date is not None:
                latest_month = latest_date.month
            
            month_count_formula = f'={latest_month}'
            month_count_cell = ws[f'A{month_count_row}']
            month_count_cell.value = month_count_formula
            month_count_cell.number_format = '0'
            month_count_cell.alignment = Alignment(horizontal='left')
            # 为month_count_formula所在单元格添加高亮处理
            month_count_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色高亮
            month_count_cell.font = Font(bold=True)  # 加粗字体

            # 先找到每年第一个月的行号，搜索所有数据行
            base_rows = {}
            for row_idx in range(4, ws.max_row + 1):
                date_value = ws[f'A{row_idx}'].value
                if date_value:
                    if isinstance(date_value, str) and 'Q' in date_value.upper():
                        continue
                    try:
                        date = pd.to_datetime(date_value)
                        if date.month == 1:  # 如果是1月
                            base_rows[date.year] = row_idx
                    except:
                        continue

            # 写入年份基准行号
            base_row_title_row = month_count_row + 1
            ws[f'A{base_row_title_row}'] = "月份"

            # 数据起始行位置调整
            ytd_data_start_row = base_row_title_row + 1
            
            # 准备年度数据
            year_groups = {}
            
            # 遍历数据处理每个月
            for _, row in data_df.iterrows():
                date = row['日期']
                year = date.year
                month = date.month
                
                # 初始化当前年度数据
                if year not in year_groups:
                    year_groups[year] = {
                        'year_label': f"{year}年",
                        'base_row': base_rows.get(year, 0),  # 获取该年份的基准行号
                        'data': {col: 0 for col in value_cols},
                        'months': set()
                    }
                
                # 记录数据
                current_group = year_groups[year]
                for col in value_cols:
                    if pd.notna(row[col]):  # 只记录非空值
                        current_group['data'][col] = row[col]  # 不需要累加，Excel公式会处理
                current_group['months'].add(month)
            
            # 按年份排序处理数据，包含所有年份（包括当前年份如2025年）
            sorted_years = sorted(year_groups.keys())
            current_row = ytd_data_start_row
            
            # 获取A列的月份设置值 (a)
            target_month_count = ws[f'A{month_count_row}'].value
            if isinstance(target_month_count, str) and target_month_count.startswith('='):
                # 如果是公式，提取公式中的数值
                try:
                    target_month_count = int(target_month_count.replace('=', ''))
                except:
                    target_month_count = None
            else:
                try:
                    target_month_count = int(float(target_month_count))
                except:
                    target_month_count = None
            
            for year in sorted_years:
                current_group = year_groups[year]
                # 获取当前年份的月份数量 (b)
                month_count = len(current_group['months'])
                
                # 写入年份标签
                ws[f'A{current_row}'] = current_group['year_label']
                
                # 为每个统计列计算YTD值
                for col in value_cols:
                    if col in ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W', 'Z', 'AC', 'AF', 'AI', 'AL', 'AO', 'AR', 'AU', 'AX']:
                        # 使用基准行号计算YTD
                        base_row = current_group['base_row']
                        # 只有当找到基准行时才计算（如果target_month_count为None则跳过月份数量检查）
                        if base_row > 0 and (target_month_count is None or month_count >= target_month_count):
                            # 对于AL、AO、AR、AU列使用AVERAGE，其他列使用SUM
                            if col in ['AL', 'AO', 'AR', 'AU']:
                                ytd_formula = f'=AVERAGE({col}{base_row}:INDEX({col}:{col}, {base_row}+$A{month_count_row}-1))'
                            else:
                                ytd_formula = f'=SUM({col}{base_row}:INDEX({col}:{col}, {base_row}+$A{month_count_row}-1))'
                            ws[f'{col}{current_row}'] = ytd_formula
                        else:
                            # 如果月份数量不足，则不计算，单元格留空
                            ws[f'{col}{current_row}'] = ""
                        
                        # 计算同比变化率
                        col_num = column_letter_to_number(col)
                        yoy_col = column_number_to_letter(col_num + 1)  # 获取下一列作为同比列
                        
                        # YTD同比上年率和同比19率公式
                        if year > 2018:  # 2018年不计算同比
                            # 计算同比上年
                            prev_year = year - 1
                            if prev_year in year_groups:
                                prev_year_base_row = year_groups[prev_year]['base_row']
                                prev_year_month_count = len(year_groups[prev_year]['months'])
                                # 只有当当前年和上一年都找到基准行时才计算同比（如果target_month_count为None则跳过月份数量检查）
                                if (prev_year_base_row > 0 and base_row > 0 and 
                                    (target_month_count is None or (month_count >= target_month_count and prev_year_month_count >= target_month_count))):
                                    # 对于AL、AO、AR、AU列使用AVERAGE，其他列使用SUM
                                    if col in ['AL', 'AO', 'AR', 'AU']:
                                        prev_formula = f'=AVERAGE({col}{prev_year_base_row}:INDEX({col}:{col}, {prev_year_base_row}+$A{month_count_row}-1))'
                                        curr_formula = f'=AVERAGE({col}{base_row}:INDEX({col}:{col}, {base_row}+$A{month_count_row}-1))'
                                    else:
                                        prev_formula = f'=SUM({col}{prev_year_base_row}:INDEX({col}:{col}, {prev_year_base_row}+$A{month_count_row}-1))'
                                        curr_formula = f'=SUM({col}{base_row}:INDEX({col}:{col}, {base_row}+$A{month_count_row}-1))'
                                    yoy_formula = f'=IF({curr_formula}="","",IF({prev_formula}="","",{prev_formula}/{curr_formula}-1))'
                                    ws[f'{yoy_col}{current_row}'] = yoy_formula
                                else:
                                    ws[f'{yoy_col}{current_row}'] = ""
                            else:
                                ws[f'{yoy_col}{current_row}'] = ""

                            # 计算同比19（简化逻辑：直接查找2019年的行）
                            if year >= 2023:
                                col_num = column_letter_to_number(col)
                                yoy19_col = column_number_to_letter(col_num + 2)  # 获取后第二列作为同比19列
                                
                                # 查找A列内容为"2019年"的行
                                row_2019 = None
                                for row_idx in range(ytd_data_start_row, current_row):
                                    if ws[f'A{row_idx}'].value == "2019年":
                                        row_2019 = row_idx
                                        break
                                
                                # 2019年数据一定存在，直接计算同比
                                if row_2019 is not None:
                                    curr_formula = f'={col}{current_row}'
                                    base_2019_formula = f'={col}{row_2019}'
                                    yoy19_formula = f'={curr_formula}/{base_2019_formula}-1'  # 直接计算当前数据/2019年数据-1
                                    ws[f'{yoy19_col}{current_row}'] = yoy19_formula
                                else:
                                    ws[f'{yoy19_col}{current_row}'] = ""
                        else:
                            ws[f'{yoy_col}{current_row}'] = ""
                
                # 添加空行（除了最后一年）
                if year != sorted_years[-1]:
                    current_row += 2
                else:
                    current_row += 1

            # 添加全年数据统计
            # 空一行，添加全年标题
            current_row += 2
            full_year_title_row = current_row
            ws[f'A{full_year_title_row}'] = "全年"
            ws[f'A{full_year_title_row}'].font = Font(bold=True)
            current_row += 1

            # 按年份统计全年数据
            yearly_totals = {}
            for row_idx in range(4, ws.max_row + 1):
                date_value = ws[f'A{row_idx}'].value
                if date_value:
                    try:
                        # 跳过季度数据
                        if isinstance(date_value, str) and ('Q' in date_value or '年' in date_value):
                            continue
                            
                        date = pd.to_datetime(date_value)
                        year = date.year
                        if year >= 2014:
                            if year not in yearly_totals:
                                yearly_totals[year] = {
                                    'months': set(),
                                    'data': {col: 0 for col in value_cols},
                                    'month_values': {col: [] for col in value_cols}  # 用于存储每月数据计算平均值
                                }
                            # 记录月份
                            yearly_totals[year]['months'].add(date.month)
                            # 累加数据（只累加月度数据）
                            for col in value_cols:
                                if col in ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W', 'Z', 'AC', 'AF', 'AI', 'AL', 'AO', 'AR', 'AU', 'AX']:
                                    cell_value = ws[f'{col}{row_idx}'].value
                                    if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                                        if col in ['AL', 'AO', 'AR', 'AU']:
                                            # 对于AL、AO、AR、AU列，存储每月的值用于计算平均值
                                            yearly_totals[year]['month_values'][col].append(cell_value)
                                        else:
                                            # 对于其他列，继续使用累加
                                            yearly_totals[year]['data'][col] = yearly_totals[year]['data'].get(col, 0) + cell_value
                    except:
                        continue

            # 按年份顺序写入全年数据
            sorted_years = sorted(yearly_totals.keys())
            for year in sorted_years:
                yearly_data = yearly_totals[year]
                # 只处理有完整12个月数据的年份
                if len(yearly_data['months']) == 12:
                    ws[f'A{current_row}'] = f"{year}年"
                    
                    for col in value_cols:
                        if col in ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W', 'Z', 'AC', 'AF', 'AI', 'AL', 'AO', 'AR', 'AU', 'AX']:
                            # 对于AL、AO、AR、AU列，计算平均值；其他列使用累加值
                            if col in ['AL', 'AO', 'AR', 'AU']:
                                if yearly_data['month_values'][col]:
                                    yearly_value = sum(yearly_data['month_values'][col]) / len(yearly_data['month_values'][col])
                                else:
                                    yearly_value = 0
                            else:
                                yearly_value = yearly_data['data'][col]
                            
                            ws[f'{col}{current_row}'] = yearly_value
                            
                            # 计算同比变化率
                            col_num = column_letter_to_number(col)
                            yoy_col = column_number_to_letter(col_num + 1)
                            
                            # 计算同比上年
                            prev_year = year - 1
                            if prev_year in yearly_totals and len(yearly_totals[prev_year]['months']) == 12:
                                # 获取上一年的值（也需要区分AL、AO、AR、AU列）
                                if col in ['AL', 'AO', 'AR', 'AU']:
                                    if yearly_totals[prev_year]['month_values'][col]:
                                        prev_value = sum(yearly_totals[prev_year]['month_values'][col]) / len(yearly_totals[prev_year]['month_values'][col])
                                    else:
                                        prev_value = 0
                                else:
                                    prev_value = yearly_totals[prev_year]['data'][col]
                                
                                if prev_value != 0:  # 避免除以零
                                    yoy_value = prev_value / yearly_value - 1
                                    ws[f'{yoy_col}{current_row}'] = yoy_value
                                    ws[f'{yoy_col}{current_row}'].number_format = '0.0%'
                            
                            # 计算同比19（仅对2023年及以后的数据）
                            if year >= 2023 and 2019 in yearly_totals and len(yearly_totals[2019]['months']) == 12:
                                yoy19_col = column_number_to_letter(col_num + 2)
                                # 获取2019年的值（也需要区分AL、AO、AR、AU列）
                                if col in ['AL', 'AO', 'AR', 'AU']:
                                    if yearly_totals[2019]['month_values'][col]:
                                        base_2019_value = sum(yearly_totals[2019]['month_values'][col]) / len(yearly_totals[2019]['month_values'][col])
                                    else:
                                        base_2019_value = 0
                                else:
                                    base_2019_value = yearly_totals[2019]['data'][col]
                                
                                if base_2019_value != 0:
                                    yoy19_value = yearly_value / base_2019_value - 1
                                    ws[f'{yoy19_col}{current_row}'] = yoy19_value
                                    ws[f'{yoy19_col}{current_row}'].number_format = '0.0%'
                    
                    current_row += 1

            return current_row

        except Exception as e:
            logger.error(f"添加YTD统计时出错: {str(e)}")
            raise

    def merge_indicators(self, indicators: List[Dict], on: str = "日期") -> pd.DataFrame:
        """
        合并多个指标数据到一个DataFrame
        
        Args:
            indicators: 指标数据字典列表
            on: 合并的键，默认为"日期"
            
        Returns:
            合并后的DataFrame
        """
        try:
            if not indicators:
                return pd.DataFrame()
            
            # 初始化结果DataFrame
            result = indicators[0]["data"].copy()
            
            # 合并其他指标
            for i in range(1, len(indicators)):
                # 获取当前指标数据
                current_data = indicators[i]["data"].copy()
                
                # 确保日期列是datetime类型
                if not pd.api.types.is_datetime64_any_dtype(result[on]):
                    result[on] = pd.to_datetime(result[on], errors='coerce')
                
                if not pd.api.types.is_datetime64_any_dtype(current_data[on]):
                    current_data[on] = pd.to_datetime(current_data[on], errors='coerce')
                
                # 合并数据
                result = pd.merge(result, current_data, on=on, how="outer")
            
            # 按日期排序
            result = result.sort_values(by=on).reset_index(drop=True)
            
            logger.info(f"已合并 {len(indicators)} 个指标数据，结果形状: {result.shape}")
            
            return result
        
        except Exception as e:
            logger.error(f"合并指标数据时出错: {str(e)}")
            raise
#同比差值
    def apply_yoy_diff_formulas(self, backup_file: str, sheet_name: str, 
                               yoy_diff_cols: List[str], start_row: int = 16) -> None:
        """
        为指定工作表的yoy_diff_cols列应用同比差值公式
        
        Args:
            backup_file: Excel文件路径
            sheet_name: 工作表名称
            yoy_diff_cols: 需要应用差值公式的列列表
            start_row: 开始行号，默认为16
        """
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(backup_file)
            ws = wb[sheet_name]
            
            # 预先收集2017年月度数据行号，用于2018年季度数据计算
            data_2017_rows = {}  # {month: row_number}
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=1).value) if ws.cell(row=row_idx, column=1).value else ""
                if cell_value.startswith("2017-"):
                    try:
                        month = int(cell_value.split("-")[1])
                        data_2017_rows[month] = row_idx
                    except (ValueError, IndexError):
                        continue
            
            # 查找YTD和全年行号，用于YTD数据特殊处理
            ytd_row = None
            full_year_row = None
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=1).value) if ws.cell(row=row_idx, column=1).value else ""
                # 检测YTD行：检查单元格的实际显示值
                if cell_value == "YTD":
                    ytd_row = row_idx
                    print(f"找到YTD行: {row_idx}, 内容: '{cell_value}'")
                elif "全年" in cell_value:
                    full_year_row = row_idx
                    print(f"找到全年行: {row_idx}, 内容: '{cell_value}'")
                    break  # 找到全年行后就可以退出了
            
            print(f"最终确定 - YTD行号: {ytd_row}, 全年行号: {full_year_row}")
            
            # 如果没有找到YTD行，尝试更广泛的搜索
            if ytd_row is None:
                print("未找到YTD行，尝试更广泛的搜索...")
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=1)
                    cell_value = str(cell.value) if cell.value else ""
                    # 检查是否包含YTD
                    if "YTD" in cell_value.upper():
                        ytd_row = row_idx
                        print(f"找到YTD行(广泛搜索): {row_idx}, 内容: '{cell_value}'")
                        break
            for col in yoy_diff_cols:
                col_num = column_letter_to_number(col)
                source_col = column_number_to_letter(col_num - 1)
                
                # 确定结束行
                last_non_empty_row = None
                for row in range(ws.max_row, start_row - 1, -1):
                    if ws.cell(row=row, column=col_num - 1).value is not None:  # 检查源数据列
                        last_non_empty_row = row
                        break
                
                end_row = last_non_empty_row if last_non_empty_row is not None else start_row
                
                # 应用公式
                for row in range(start_row, end_row + 1):
                    # 检查被减数（当前值）是否为空
                    current_cell = ws.cell(row=row, column=col_num - 1)
                    if current_cell.value is None or current_cell.value == "":
                        # 如果被减数为空，清除单元格
                        cell = ws.cell(row=row, column=col_num)
                        cell.value = None
                        continue
                    
                    # 获取当前行A列的值
                    a_cell_value = str(ws.cell(row=row, column=1).value) if ws.cell(row=row, column=1).value else ""
                    
                    # 检查是否为2018年季度数据（A列同时包含"Q"和"18"）
                    if "Q" in a_cell_value and "18" in a_cell_value:
                        # 解析季度信息
                        try:
                            quarter_num = int(a_cell_value.split("Q")[1])
                            
                            # 根据季度确定对应的2017年月份
                            if quarter_num == 1:
                                months_2017 = [1, 2, 3]  # Q1: 1-3月
                            elif quarter_num == 2:
                                months_2017 = [4, 5, 6]  # Q2: 4-6月
                            elif quarter_num == 3:
                                months_2017 = [7, 8, 9]  # Q3: 7-9月
                            elif quarter_num == 4:
                                months_2017 = [10, 11, 12]  # Q4: 10-12月
                            else:
                                continue
                            #季度数据同比差值的特殊情况
                            # 构建2017年季度平均值的公式
                            valid_rows_2017 = []
                            for month in months_2017:
                                if month in data_2017_rows:
                                    valid_rows_2017.append(data_2017_rows[month])
                            
                            if len(valid_rows_2017) == 3:  # 只有当3个月数据都存在时才计算
                                # 构建平均值公式：AVERAGE(2017年Q对应月份的数据)
                                avg_formula_parts = [f"{source_col}{row_2017}" for row_2017 in valid_rows_2017]
                                avg_formula = f"AVERAGE({','.join(avg_formula_parts)})"
                                
                                # 构建差值公式：当前季度数据 - 2017年对应季度平均值
                                formula = f"={source_col}{row}-({avg_formula})"
                                cell = ws.cell(row=row, column=col_num)
                                cell.value = formula
                                cell.number_format = "0.00"
                            else:
                                # 如果2017年数据不完整，清除单元格
                                cell = ws.cell(row=row, column=col_num)
                                cell.value = None
                        except (ValueError, IndexError):
                            # 如果解析季度失败，使用默认逻辑
                            prev_row = row - 12
                            if prev_row >= 1:
                                formula = f"={source_col}{row}-{source_col}{prev_row}"
                                cell = ws.cell(row=row, column=col_num)
                                cell.value = formula
                                cell.number_format = "0.00"
                    else:
                        # 检查是否为其他年份的季度数据（A列含有"Q"但不含"18"）
                        if "Q" in a_cell_value and "18" not in a_cell_value:
                            # 解析季度信息，查找上一年同期季度数据
                            try:
                                # 提取年份和季度
                                if "Q" in a_cell_value:
                                    year_part = a_cell_value.split("Q")[0]
                                    quarter_part = a_cell_value.split("Q")[1]
                                    
                                    # 处理两位年份格式（如19Q1）
                                    if len(year_part) == 2:
                                        current_year = int("20" + year_part)
                                    else:
                                        current_year = int(year_part)
                                    
                                    quarter_num = int(quarter_part)
                                    prev_year = current_year - 1
                                    
                                    # 构建上一年同期季度标签
                                    prev_year_short = str(prev_year)[-2:]  # 取后两位
                                    prev_quarter_label = f"{prev_year_short}Q{quarter_num}"
                                    
                                    # 查找上一年同期季度数据行
                                    prev_quarter_row = None
                                    for search_row in range(start_row, ws.max_row + 1):
                                        search_cell_value = str(ws.cell(row=search_row, column=1).value) if ws.cell(row=search_row, column=1).value else ""
                                        if search_cell_value == prev_quarter_label:
                                            prev_quarter_row = search_row
                                            break
                                    
                                    if prev_quarter_row is not None:
                                        # 构建季度同比差值公式：本期数据 - 上一年同期数据
                                        formula = f"={source_col}{row}-{source_col}{prev_quarter_row}"
                                        cell = ws.cell(row=row, column=col_num)
                                        cell.value = formula
                                        cell.number_format = "0.00"
                                    else:
                                        # 如果找不到上一年同期数据，清除单元格
                                        cell = ws.cell(row=row, column=col_num)
                                        cell.value = None
                                else:
                                    # 如果解析失败，使用默认月度逻辑
                                    prev_row = row - 12
                                    if prev_row >= 1:
                                        formula = f"={source_col}{row}-{source_col}{prev_row}"
                                        cell = ws.cell(row=row, column=col_num)
                                        cell.value = formula
                                        cell.number_format = "0.00"
                            except (ValueError, IndexError):
                                # 如果解析失败，使用默认月度逻辑
                                prev_row = row - 12
                                if prev_row >= 1:
                                    formula = f"={source_col}{row}-{source_col}{prev_row}"
                                    cell = ws.cell(row=row, column=col_num)
                                    cell.value = formula
                                    cell.number_format = "0.00"
                        else:
                            # 对于非季度数据，检查是否为YTD数据的特殊情况或全年数据的特殊情况
                            # 检查当前行是否在YTD和全年之间且A列含有"年"
                            is_ytd_special_case = False
                            is_full_year_special_case = False
                            
                            if (ytd_row is not None and full_year_row is not None and 
                                ytd_row < row < full_year_row and "年" in a_cell_value):
                                is_ytd_special_case = True
                                print(f"检测到YTD特殊情况 - 行{row}: '{a_cell_value}', YTD行: {ytd_row}, 全年行: {full_year_row}")
                            elif (full_year_row is not None and row > full_year_row and "年" in a_cell_value):
                                is_full_year_special_case = True
                                print(f"检测到全年数据特殊情况 - 行{row}: '{a_cell_value}', 全年行: {full_year_row}")
                            
                            if is_ytd_special_case:
                                # YTD数据特殊处理：prev_row = row - 2
                                prev_row = row - 2
                                if prev_row >= 1:
                                    formula = f"={source_col}{row}-{source_col}{prev_row}"
                                    cell = ws.cell(row=row, column=col_num)
                                    cell.value = formula
                                    cell.number_format = "0.0"  # YTD数据保留1位小数
                                    print(f"应用YTD特殊公式到行{row}: {formula}")
                            elif is_full_year_special_case:
                                # 全年数据特殊处理：当前行数据-上一行数据 (period=1)，从2016年开始计算
                                try:
                                    year = int(a_cell_value.replace("年", ""))
                                    if year >= 2016:  # 从2016年开始计算全年数据的同比差值
                                        prev_row = row - 1
                                        if prev_row >= 1:
                                            formula = f"={source_col}{row}-{source_col}{prev_row}"
                                            cell = ws.cell(row=row, column=col_num)
                                            cell.value = formula
                                            cell.number_format = "0.0"  # 全年数据保留1位小数
                                            print(f"应用全年数据特殊公式到行{row}: {formula}")
                                    else:
                                        # 2016年之前的全年数据不计算同比差值
                                        cell = ws.cell(row=row, column=col_num)
                                        cell.value = None
                                        print(f"跳过{year}年全年数据同比差值计算")
                                except (ValueError, IndexError):
                                    # 如果无法解析年份，使用默认逻辑
                                    prev_row = row - 1
                                    if prev_row >= 1:
                                        formula = f"={source_col}{row}-{source_col}{prev_row}"
                                        cell = ws.cell(row=row, column=col_num)
                                        cell.value = formula
                                        cell.number_format = "0.0"  # 全年数据保留1位小数
                                        print(f"应用全年数据特殊公式到行{row}: {formula}")
                            else:
                                # 对于其他非季度数据，使用原有的月度同比差值逻辑
                                prev_row = row - 12
                                if prev_row >= 1:
                                    formula = f"={source_col}{row}-{source_col}{prev_row}"
                                    cell = ws.cell(row=row, column=col_num)
                                    cell.value = formula
                                    cell.number_format = "0.00"
            
            wb.save(backup_file)
            logger.info(f"已为 {sheet_name} 工作表的yoy_diff_cols列应用差值公式（包含2018年季度特殊处理）")
            
        except Exception as e:
            logger.error(f"应用yoy_diff_cols公式时出错: {str(e)}")
            raise
            
  #差值型同比19
    def apply_diff19_formulas(self, backup_file: str, sheet_name: str, 
                             diff19_cols: List[str], start_row: int = 112) -> None:
        """
        为指定工作表的diff19_cols列应用与2019年同期的差值公式
        
        Args:
            backup_file: Excel文件路径
            sheet_name: 工作表名称
            diff19_cols: 需要应用差值公式的列列表
            start_row: 开始行号，默认为112
        """
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(backup_file)
            ws = wb[sheet_name]
            
            # 提前找到2019年各季度和年度数据所在的行
            row_2019Q1 = None
            row_2019Q2 = None
            row_2019Q3 = None
            row_2019Q4 = None
            row_2019 = None
            full_year_row = None  # 查找"全年"所在的行
            row_2019_after_full_year = None  # 查找"全年"行之后的"2019年"行
            ytd_row = None  # 查找YTD行
            
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=1).value)
                if "19Q1" in cell_value:
                    row_2019Q1 = row_idx
                elif "19Q2" in cell_value:
                    row_2019Q2 = row_idx
                elif "19Q3" in cell_value:
                    row_2019Q3 = row_idx
                elif "19Q4" in cell_value:
                    row_2019Q4 = row_idx
                elif cell_value == "2019年":
                    row_2019 = row_idx
                elif "全年" in cell_value:
                    full_year_row = row_idx
                elif cell_value == "YTD" or "YTD" in cell_value.upper():
                    ytd_row = row_idx
                    full_year_row = row_idx
                if all(row is not None for row in [row_2019Q1, row_2019Q2, row_2019Q3, row_2019Q4, row_2019]):
                    break
            
            # 查找"全年"行之后的特殊行
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=1).value)
                if "全年" in cell_value:
                    full_year_row = row_idx
                    break
            
            # 如果找到了"全年"行，查找其后的"2019年"行
            if full_year_row is not None:
                for row_idx in range(full_year_row + 1, ws.max_row + 1):
                    cell_value = str(ws.cell(row=row_idx, column=1).value)
                    if cell_value == "2019年":
                        row_2019_after_full_year = row_idx
                        break
            
            for col in diff19_cols:
                col_num = column_letter_to_number(col)
                source_col = column_number_to_letter(col_num - 2)
                
                # 确定结束行
                last_non_empty_row = None
                for row in range(ws.max_row, start_row - 1, -1):
                    if ws.cell(row=row, column=col_num - 2).value is not None:  # 检查源数据列
                        last_non_empty_row = row
                        break
                
                end_row = last_non_empty_row if last_non_empty_row is not None else start_row
                
                # 应用公式
                for row in range(start_row, end_row + 1):
                    date_cell = ws.cell(row=row, column=1)
                    date_value = str(date_cell.value) if date_cell.value else ""
                    formula = ""
                    
                    try:
                        # 处理月度数据（格式：yyyy-mm）
                        if "-" in date_value and not "Q" in date_value:
                            year = int(date_value.split("-")[0])
                            month = date_value.split("-")[1]
                            
                            # 只对112行及之后的数据进行同比19计算
                            if row >= 112:
                                # 寻找2019年同月的数据行
                                month_19_row = None
                                for search_row in range(1, ws.max_row + 1):
                                    search_date = str(ws.cell(row=search_row, column=1).value)
                                    # 确保匹配的是2019年且月份完全相同
                                    if (search_date.startswith("2019-") and 
                                        search_date.split("-")[1] == month):
                                        month_19_row = search_row
                                        break
                            
                            # 对112行及之后的月份数据计算与2019年同期的差值
                            if row >= 112 and month_19_row:  # 如果是112行之后且找到了对应的2019年同月数据
                                formula = f"={source_col}{row}-{source_col}{month_19_row}"
                        
                        # 处理季度数据
                        elif "Q" in date_value:
                            year = int(date_value[:date_value.find("Q")])
                            quarter = date_value[date_value.find("Q")+1]
                            
                            if year >= 23:  # 只处理23年及以后的数据
                                # 根据季度找到对应的2019年季度数据行
                                row_2019_q = None
                                if quarter == "1" and row_2019Q1 is not None:
                                    row_2019_q = row_2019Q1
                                elif quarter == "2" and row_2019Q2 is not None:
                                    row_2019_q = row_2019Q2
                                elif quarter == "3" and row_2019Q3 is not None:
                                    row_2019_q = row_2019Q3
                                elif quarter == "4" and row_2019Q4 is not None:
                                    row_2019_q = row_2019Q4
                                    
                                if row_2019_q is not None:
                                    formula = f"={source_col}{row}-{source_col}{row_2019_q}"
                                
                        # 处理2023年和2024年年度数据与2019年的差值
                        elif "年" in date_value:
                            year = int(date_value.replace("年", ""))
                            if year >= 2023:
                                # 检查是否在"全年"行之后
                                if full_year_row is not None and row > full_year_row:
                                    # 新的特殊处理：使用"全年"行之后的"2019年"行作为基准
                                    if row_2019_after_full_year is not None:
                                        formula = f"={source_col}{row}-{source_col}{row_2019_after_full_year}"
                                else:
                                    # 原有的处理逻辑：使用普通的"2019年"行
                                    if row_2019 is not None:
                                        formula = f"={source_col}{row}-{source_col}{row_2019}"
                    except (ValueError, IndexError):
                        pass
                    
                    # 写入公式
                    if formula:
                        cell = ws.cell(row=row, column=col_num)
                        cell.value = formula
                        
                        # 根据数据类型设置不同的小数位数
                        if (ytd_row is not None and full_year_row is not None and 
                            ytd_row < row < full_year_row and "年" in date_value):
                            # YTD数据保留1位小数
                            cell.number_format = "0.0"
                        elif (full_year_row is not None and row > full_year_row and "年" in date_value):
                            # 全年数据保留1位小数
                            cell.number_format = "0.0"
                        else:
                            # 其他数据保留2位小数
                            cell.number_format = "0.00"
            
            wb.save(backup_file)
            logger.info(f"已为 {sheet_name} 工作表的diff19_cols列应用差值公式")
            
        except Exception as e:
            logger.error(f"应用diff19_cols公式时出错: {str(e)}")
            raise

