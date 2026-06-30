import pandas as pd
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class PublicPlugin:
    @staticmethod
    def public_write_data(context, params):
        """
        功能说明：
            处理公用事业基础指标数据，写入到目标工作表中。
            从 data_reader 缓存中读取多个指标数据，以第一个指标的日期为基准，
            将所有指标数据按此日期对齐写入 Excel。
            支持单位转换（从指定行读取转换因子），日期按降序排列（最新日期在前）。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - start_row: int (默认 50) — 数据写入的起始行号。
                - start_column: int (默认 1) — 日期列写入的起始列号。
                - unit_conversion: int (默认 0) — 单位转换因子所在的行号。
                - start_date: str (默认 '2014-01-01') — 数据最早日期。
                - date_format: str (默认 'yyyy-mm') — 日期列的数字格式。
        """ 
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 50)
        start_column = params.get('start_column', 1)
        unit_conversion_row = params.get('unit_conversion', 0)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm')
        start_date_ts = pd.to_datetime(start_date)

        def convert_value(value, unit_conversion):
            """根据列和配置的倍率转换数值"""
            if value is None or not isinstance(value, (int, float)):
                return value
            if unit_conversion:
                return value * unit_conversion
            return value
                
        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=False).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表（不再补全日期）
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        # 直接使用第一个指标的数据作为基础日期表
        base_df = first_df.copy()

        # 3. 写入统一日期列和指标数据（应用单位转换）
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            # 第1列写日期
            date_cell = ws.cell(row=current_row, column=start_column, value=current_date)
            date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))
            
            # 添加安全检查
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
                    
                if unit_conversion_row >= 1 and col_num >= 1:
                    unit_cell = ws.cell(row=unit_conversion_row, column=col_num)
                    if unit_cell.value is not None and unit_cell.value > 0:
                        unit_conversion = unit_cell.value
                    else:
                        unit_conversion = 1
                else:
                    unit_conversion = 1
            except Exception as e:
                print(f"    - 获取单位转换因子失败，使用默认值1: {e}")
                unit_conversion = 1
                
            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                converted_value = convert_value(value, unit_conversion)
                ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=converted_value
                )
        
        print(f"    - 完成公用事业基础数据写入")
    
    @staticmethod
    def public_elec_write_title(context, params):
        """
        功能说明：
            写入公用事业（电力）报表的动态标题和表头日期。
            根据最新日期（latest_row 行 A 列的值）自动计算并格式化标题，
            支持主标题和副标题两行，分别使用当前月份和上月（或上一年12月）的数据。
            特殊处理 2 月份：标题显示为 "1-2" 表示 1-2 月合并数据。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - latest_row: int — 最新数据所在行号，用于获取最新日期。
                - title1_row: int — 主标题所在行号。
                - title1_col: str — 主标题所在列字母。
                - title2_row: int — 副标题所在行号（用于上月/上期数据）。
                - title2_col: str — 副标题所在列字母。
        """
        ws = context['ws']

        latest_row = params.get('latest_row')
        latest_date = ws.cell(row=latest_row, column=1).value
        latest_year = latest_date.year
        latest_month = latest_date.month

        def write_title(title_row, title_col, year, month):
            if month == 2:
                month = "1-2"
            # 写入主标题
            title_template = ws.cell(row=title_row, column=title_col).value
            title = title_template.format(year=year, month=month)
            ws.cell(row=title_row, column=title_col, value=title)
            
            header_row = title_row + 1
            hearer_col = title_col + 1
            # 写入表头日期1
            header_template = ws.cell(row=header_row, column=hearer_col).value
            header = header_template.format(year=year, month=month)
            ws.cell(row=header_row, column=hearer_col, value=header)
            # 写入表头日期2
            header_col = hearer_col + 2
            year = year - 1
            header_template = ws.cell(row=header_row, column=header_col).value
            header = header_template.format(year=year, month=month)
            ws.cell(row=header_row, column=header_col, value=header)

        title1_row = params.get('title1_row')
        title1_col = column_letter_to_number(params.get('title1_col'))
        write_title(title1_row, title1_col, latest_year, latest_month)
        if latest_month == 2:
            month2 = 12
            year2 = latest_year - 1
        else:
            month2 = latest_month - 1
            year2 = latest_year
        title2_row = params.get('title2_row')
        title2_col = column_letter_to_number(params.get('title2_col'))
        write_title(title2_row, title2_col, year2, month2)

    @staticmethod
    def public_t3_write_header(context, params):
        """
        功能说明：
            写入 T+3（三日滚动）报表的表头日期。
            根据最新日期（latest_row 行 A 列的值）和其下一行日期，
            分别写入两个表头单元格的日期（月/日格式）。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - latest_row: int — 最新数据所在行号，用于获取最新日期。
                - header_row: int — 表头所在行号。
               
        """        
        ws = context['ws']

        def write_header(header_row, header_col, month, day):
            header_template = ws.cell(row=header_row, column=header_col).value
            header = header_template.format(month=month, day=day)
            ws.cell(row=header_row, column=header_col, value=header)
        
        latest_row = params.get('latest_row')
        latest_date = ws.cell(row=latest_row, column=1).value
        latest_month = latest_date.month
        latest_day = latest_date.day
        header_row = params.get('header_row')
        header1_col = column_letter_to_number(params.get('header_col'))
        write_header(header_row, header1_col, latest_month, latest_day)
        
        header2_date = ws.cell(row=latest_row+1, column =1).value
        header2_month = header2_date.month
        header2_day = header2_date.day
        header2_col = header1_col + 1
        write_header(header_row, header2_col, header2_month, header2_day)
        