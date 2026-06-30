import pandas as pd
import re
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.datetime import from_excel as datetime_from_excel

class MedicalPlugin:
    """
    医疗健康行业数据插件，支持横向（宽表）数据写入:

    源数据格式（纵向）:
        日期       | 指标值
    2020-01-01    | 6.36
    2020-02-01    | 2.56

    目标模板写入后格式（横向）:
        A列: 指标名称 | B列起: 2020.1 | 2020.2 | ...
        第4行: 国内医疗健康融资 | 6.36   | 2.56   | ...
    """

    @staticmethod
    def medical_write_data(context, params):
        """
        横向写入医疗数据：
        1. 从所有指标的源数据中提取统一的时间周期序列
        2. 将时间周期作为列头写入目标模板
        3. 将每个指标的数据按日期对号写入对应列

        参数:
            date_header_row: 日期列头写入行号 (默认1)
            data_start_column: 数据起始列号 (默认2)
            start_date: 数据起始日期，此前的数据将被过滤
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_row_map = sheet_config['indicators']  # code -> target_row

        date_header_row = params.get('date_header_row', 1)
        data_start_column = params.get('data_start_column', 2)
        start_date = params.get('start_date', '2020-01-01')
        start_date_ts = pd.to_datetime(start_date)

        # ---------------------------------------------------------------
        # 1. 读取所有指标数据，提取统一的时间周期序列
        # ---------------------------------------------------------------
        indicator_dfs = {}
        all_periods = set()

        for indicator_code in indicator_row_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator_code)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                print(f"    - 无数据，跳过: {indicator_code}")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()

            if df.empty:
                print(f"    - 无有效数据: {indicator_code}")
                continue

            indicator_dfs[indicator_code] = df

            for _, row in df.iterrows():
                all_periods.add((row['日期'].year, row['日期'].month))

        if not all_periods:
            print(f"    - 所有指标均无有效数据，跳过写入")
            return

        sorted_periods = sorted(all_periods)  # [(2020,1), (2020,2), ...]
        print(f"    - 共 {len(sorted_periods)} 个时间周期, 从 {sorted_periods[0][0]}.{sorted_periods[0][1]} 到 {sorted_periods[-1][0]}.{sorted_periods[-1][1]}")

        # ---------------------------------------------------------------
        # 2. 写入日期列头
        # ---------------------------------------------------------------
        MedicalPlugin._unmerge_row(ws, date_header_row)
        for i, (year, month) in enumerate(sorted_periods):
            col = data_start_column + i
            ws.cell(row=date_header_row, column=col, value=f"{year}.{month}")

        # ---------------------------------------------------------------
        # 3. 逐一写入指标数据
        # ---------------------------------------------------------------
        for indicator_code, target_row in indicator_row_map.items():
            if indicator_code not in indicator_dfs:
                continue

            df = indicator_dfs[indicator_code]
            value_col = df.columns[-1]

            # 构建 (year,month) -> value 映射
            value_map = {}
            for _, row in df.iterrows():
                value_map[(row['日期'].year, row['日期'].month)] = row[value_col]

            # 解除目标行合并单元格
            MedicalPlugin._unmerge_row(ws, target_row)

            write_count = 0
            for i, (year, month) in enumerate(sorted_periods):
                key = (year, month)
                if key in value_map:
                    ws.cell(row=target_row, column=data_start_column + i, value=value_map[key])
                    write_count += 1

            print(f"    - 指标 {indicator_code} -> 第{target_row}行, 写入 {write_count} 个数据点")

        print(f"    - 完成医疗数据写入")

    @staticmethod
    def _unmerge_row(ws, target_row):
        """解除覆盖目标行的所有合并单元格，以便写入数据"""
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= target_row <= merged_range.max_row:
                ws.unmerge_cells(str(merged_range))

    @staticmethod
    def _parse_period(period_str: str):
        """将 '2020.1' / '2020.01' 解析为 (year, month)"""
        match = re.match(r'(\d{4})[.](\d{1,2})', period_str)
        if match:
            return int(match.group(1)), int(match.group(2))
        return None


    @staticmethod
    def medical_write_quarter(context, params):
        """
        根据date_row中的月度数据自动判断所属季度，在quarter_row中合并单元格并写入yyyyQn

        参数:
            quarter_row: 季度信息写入行号 (默认2)
            date_row: 月度日期所在行号 (默认3)
            data_start_column: 数据起始列号 (默认2)
        """
        ws = context["ws"]
        quarter_row = params.get("quarter_row", 2)
        date_row = params.get("date_row", 3)
        data_start_column = params.get("data_start_column", 2)

        # 1. 扫描 date_row，读取所有列头并解析为 (col, year, month)
        periods = []
        col = data_start_column
        while True:
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value is None:
                break
            parsed = MedicalPlugin._parse_period(str(cell_value))
            if parsed:
                periods.append((col, parsed[0], parsed[1]))
            col += 1

        if not periods:
            print("    - date_row 中未找到日期数据，跳过季度写入")
            return

        # 2. 将连续的同一季度月份分组
        quarter_groups = []
        current_group = None

        for period_col, year, month in periods:
            q = (month - 1) // 3 + 1
            if (current_group is None
                    or current_group[0] != year
                    or current_group[1] != q
                    or period_col != current_group[3] + 1):
                if current_group is not None:
                    quarter_groups.append(current_group)
                current_group = [year, q, period_col, period_col]
            else:
                current_group[3] = period_col

        if current_group is not None:
            quarter_groups.append(current_group)

        # 3. 解除 quarter_row 合并，写入季度
        MedicalPlugin._unmerge_row(ws, quarter_row)
        for year, q, start_col, end_col in quarter_groups:
            cell = ws.cell(row=quarter_row, column=start_col)
            cell.value = f"{year}Q{q}"
            if start_col != end_col:
                ws.merge_cells(
                    start_row=quarter_row, start_column=start_col,
                    end_row=quarter_row, end_column=end_col
                )
        print(f"    - 完成季度行写入: 共 {len(quarter_groups)} 个季度")
    
    @staticmethod
    def medical_write_formula(context, params):
        """
        功能说明：
            从起始列开始，横向向右写入公式，直到日期行中最后一个非空列为止。

        支持的占位符：
            {col}        -> 当前列字母（A、B、C...）
            {col-N}      -> 当前列向左偏移 N 位的列字母
            {row}        -> 当前行号
            {source_row} -> 源数据行号（从 params 中获取，如 4、7、10...）
        """
        ws = context['ws']
        start_row = params.get('start_row')
        start_column = params.get('start_column', 2)
        date_row = params.get('date_row')
        formula_template = params.get('target_formula', '')
        custom_format = params.get('format', None)
        source_row = params.get('source_row', None)

        if not formula_template or not start_row or not date_row:
            print("    - missing required params, skip formula write")
            return

        # Determine end column by scanning date_row
        col = start_column
        end_column = start_column - 1
        while True:
            if ws.cell(row=date_row, column=col).value is None:
                end_column = col - 1
                break
            col += 1

        if end_column < start_column:
            print("    - no data in date_row, skip formula write")
            return

        # Find the max offset in {col-N} to skip invalid columns
        offset_pat = re.compile(r'\{col-(\d+)\}')
        max_offset = 0
        for m in offset_pat.finditer(formula_template):
            max_offset = max(max_offset, int(m.group(1)))

        write_count = 0
        for current_col in range(start_column, end_column + 1):
            if max_offset > 0 and current_col - max_offset < start_column:
                continue

            formula = formula_template

            formula = offset_pat.sub(
                lambda m: get_column_letter(current_col - int(m.group(1))),
                formula
            )
            formula = formula.replace('{col}', get_column_letter(current_col))
            formula = formula.replace('{row}', str(start_row))
            if source_row is not None:
                formula = formula.replace('{source_row}', str(source_row))

            cell = ws.cell(row=start_row, column=current_col)
            if isinstance(cell, MergedCell):
                # skip cells that are part of a merged range
                continue
            cell.value = formula
            if custom_format:
                cell.number_format = custom_format
            write_count += 1

        end_letter = get_column_letter(end_column)
        print(f"    - wrote {write_count} formulas, row {start_row}, cols {get_column_letter(max(start_column, start_column+max_offset))}-{end_letter}")

    @staticmethod
    def medical_merge_quarter_data(context, params):
        """
        合并季度数据：根据date_row中的月份，将每个季度的数据合并到该季度的起始列

        参数:
            rows: 需要处理的数据行号列表 (默认 [13,14,15])
            start_column: 数据起始列号 (默认2)
            date_row: 月度日期所在行号 (默认3)
        """
        ws = context["ws"]
        rows = params.get("rows", [13, 14, 15])
        start_column = params.get("start_column", 2)
        date_row = params.get("date_row", 3)
        
        # 1. 扫描 date_row，读取所有列头并解析为 (col, year, month)
        periods = []
        col = start_column
        while True:
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value is None:
                break
            parsed = MedicalPlugin._parse_period(str(cell_value))
            if parsed:
                periods.append((col, parsed[0], parsed[1]))
            col += 1
        
        if not periods:
            print("    - date_row 中未找到日期数据，跳过季度数据合并")
            return
        
        # 2. 按季度分组，注意跨年情况
        quarter_groups = []
        current_group = None
        
        for period_col, year, month in periods:
            # 计算季度 (1-3月为Q1, 4-6月为Q2, 7-9月为Q3, 10-12月为Q4)
            q = (month - 1) // 3 + 1
            
            # 判断是否属于同一个季度组
            is_same_quarter = False
            if current_group is not None:
                # 检查是否同一季度（考虑跨年：如2025年12月和2026年1月不在同一季度）
                if current_group[0] == year and current_group[1] == q:
                    # 检查列是否连续
                    if period_col == current_group[3] + 1:
                        is_same_quarter = True
            
            if not is_same_quarter:
                # 开始新的季度组
                if current_group is not None:
                    quarter_groups.append(current_group)
                current_group = [year, q, period_col, period_col]  # [year, quarter, start_col, end_col]
            else:
                # 更新当前组的结束列
                current_group[3] = period_col
        
        if current_group is not None:
            quarter_groups.append(current_group)
        
        # 3. 对每个季度组，合并数据行
        for year, q, start_col, end_col in quarter_groups:
            # 获取该季度的所有月份列
            quarter_cols = list(range(start_col, end_col + 1))
            
            # 遍历指定的数据行
            for row in rows:
                
                # 对于该季度组，找到有值的月份列（优先取季度末月份的值）
                # 季度末月份：3月(Q1), 6月(Q2), 9月(Q3), 12月(Q4)
                quarter_end_month = q * 3
                
                # 查找季度末月份对应的列
                end_month_col = None
                for col in quarter_cols:
                    # 从periods中查找该列对应的月份
                    for p_col, p_year, p_month in periods:
                        if p_col == col and p_month == quarter_end_month:
                            end_month_col = col
                            break
                    if end_month_col:
                        break
                
                # 如果找到了季度末月份列，使用该列的值
                if end_month_col:
                    value = ws.cell(row=row, column=end_month_col).value
                    value_col = end_month_col
                else:
                    # 如果没有季度末月份，使用该季度最后一列的值
                    value = ws.cell(row=row, column=end_col).value
                    value_col = end_col
                
                # 如果该行在季度范围内有数据，但没有找到合适的值，尝试获取第一个有值的列
                if value is None:
                    for col in quarter_cols:
                        val = ws.cell(row=row, column=col).value
                        if val is not None:
                            value = val
                            break
                
                # 先清除该季度范围内所有列的数据（在合并之前）
                for col in quarter_cols:
                    # 检查单元格是否可写（不是已合并的只读单元格）
                    cell = ws.cell(row=row, column=col)
                    # 尝试读取坐标，如果是MergedCell，会抛出AttributeError
                    try:
                        # 检查是否是合并单元格的一部分
                        cell.value = None
                    except AttributeError:
                        # 如果是MergedCell，跳过
                        pass
                
                # 写入合并后的值到季度起始列，使用常规数字格式
                ws.cell(row=row, column=start_col).value = value
                ws.cell(row=row, column=start_col).number_format = ws.cell(row=row, column=value_col).number_format
                
                # 如果该季度有多列，合并单元格
                if start_col != end_col:
                    # 合并该行从start_col到end_col的单元格
                    ws.merge_cells(
                        start_row=row, start_column=start_col,
                        end_row=row, end_column=end_col
                    )
                print(f"    - 合并数据行: {row}")
            
            # 清除该季度其他列的数据（除起始列外），但要排除已合并的只读单元格
            for col in range(start_col + 1, end_col + 1):
                for r in rows:
                    try:
                        ws.cell(row=r, column=col).value = None
                    except AttributeError:
                        # 如果是MergedCell，跳过
                        pass
        
        print(f"    - 完成季度数据合并: 共 {len(quarter_groups)} 个季度")



