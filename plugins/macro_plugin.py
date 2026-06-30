import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime, timedelta
from zhdate import ZhDate


class MacroPlugin:
    
    @staticmethod
    def macro_write_indicator_group(context, config):
        """
        功能说明：
            写入一组指标数据及参考日期列到目标工作表。
            以第一个指标的日期为基准日期列，后续指标按相同行数依次写入相邻列。
            支持数据排序和日期格式自定义。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            config: dict — 宏指标配置，包含：
                - indicators: list[str] — 指标名称列表。
                - start_row: int — 数据写入的起始行号。
                - start_col: str | int — 数据写入的起始列（列字母或列号）。
                - date_format: str (默认 'yyyy-mm-dd') — 日期列的数字格式。
                - ascending: bool (默认 False) — 日期排序方式，True 为升序，False 为降序。
        """
        ws = context['ws']
        reader = context['data_reader']
        source_sheet = context['sheet_config']['source_sheet']
        indicators = config['indicators']
        start_row = config['start_row']
        start_col = config['start_col']
        date_format = config.get('date_format', 'yyyy-mm-dd')
        ascending = config.get('ascending', False)
        
        start_col_num = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
        
        reference_dates = None
        
        from core_engine.data_processor import DataProcessor
        processor = DataProcessor(reader)
        
        for i, indicator in enumerate(indicators):
            indicator_data = reader.read_indicator_data(source_sheet, indicator)
            sorted_data = processor.sort_by_date(indicator_data, ascending=ascending)
            df = sorted_data["data"]
            
            if i == 0:
                reference_dates = df['日期'].tolist()
                for row_idx, row in df.iterrows():
                    target_row = start_row + row_idx
                    date_val = row['日期']
                    val = row[df.columns[-1]]
                    if val == 0: val = None
                    
                    ws.cell(row=target_row, column=start_col_num, value=date_val).number_format = date_format
                    ws.cell(row=target_row, column=start_col_num + 1, value=val)
                print(f"[{ws.title}] 写入参考日期及指标1: {indicator} 到 {start_col}")
            else:
                target_col = start_col_num + i + 1
                for row_idx, row in df.iterrows():
                    target_row = start_row + row_idx
                    val = row[df.columns[-1]]
                    if val == 0: val = None
                    ws.cell(row=target_row, column=target_col, value=val)
                print(f"[{ws.title}] 写入指标{i+1}: {indicator} 到列 {get_column_letter(target_col)}")

    @staticmethod
    def macro_create_pivot_table(context, config):
        """
        功能说明：
            创建基础透视表，将日期数据按"月-日"为行、年份为列进行透视。
            自动处理闰年 2月29日 到 2月28日 的映射，并支持限定最近 N 年的数据。

        params:
            context: dict — Pipeline 上下文，包含 ws、data_reader、sheet_config。
            config: dict — 透视表配置，包含：
                - indicator_code: str — 指标代码。
                - start_row: int — 透视表写入的起始行号。
                - start_col: str | int — 透视表写入的起始列（列字母或列号）。
                - years: int (默认 7) — 包含的最近年份数量。
        """
        ws = context['ws']
        reader = context['data_reader']
        source_sheet = context['sheet_config']['source_sheet']
        indicator_code = config['indicator_code']
        start_row = config['start_row']
        start_col = config['start_col']
        years = config.get('years', 7)
        
        start_col_num = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
        
        from core_engine.data_processor import DataProcessor
        processor = DataProcessor(reader)
        
        indicator_data = reader.read_indicator_data(source_sheet, indicator_code)
        sorted_data = processor.sort_by_date(indicator_data, ascending=True)
        df = sorted_data["data"].copy()
        
        df['日期'] = pd.to_datetime(df['日期'])
        df['年份'] = df['日期'].dt.year
        df['月日'] = df['日期'].dt.strftime('%m-%d')
        df['月日'] = df['月日'].replace('02-29', '02-28')
        
        current_year = datetime.now().year
        start_year = current_year - years + 1
        
        df_filtered = df[df['年份'] >= start_year].copy()
        df_filtered = df_filtered.drop_duplicates(subset=['月日', '年份'])
        pivot_df = df_filtered.pivot(index='月日', columns='年份', values=df.columns[-3])
        
        # 写入表头
        for i, year in enumerate(pivot_df.columns):
            ws.cell(row=start_row, column=start_col_num + i + 1, value=year)
            
        # 写入行标题和数据
        for i, (month_day, row_data) in enumerate(pivot_df.iterrows()):
            try:
                excel_date = datetime.strptime(f"1900-{month_day}", "%Y-%m-%d")
            except Exception:
                excel_date = None
                
            ws.cell(row=start_row + i + 1, column=start_col_num, value=excel_date)
            if excel_date:
                ws.cell(row=start_row + i + 1, column=start_col_num).number_format = "mm-dd"
                
            for j, year in enumerate(pivot_df.columns):
                value = row_data.get(year, None)
                if pd.isna(value) or value == 0:
                    value = None
                ws.cell(row=start_row + i + 1, column=start_col_num + j + 1, value=value)
                
        print(f"[{ws.title}] 创建透视表 {indicator_code} 到 {start_col}")

    # ==================================
    # 农历与春节相关逻辑
    # ==================================
    @staticmethod
    def add_global_spring_festival_tags(df: pd.DataFrame):
        """
        功能说明：
            为包含农历日期和星期的 DataFrame 添加"春节标签"列。
            根据农历"正月初一"定位春节日期，计算每一行相对于春节的周数位置，
            生成"节前N周"、"春节当周"、"节后N周"等标签。

        params:
            df: pd.DataFrame — 必须包含以下列：
                - 农历: str — 农历日期字符串（如"正月初一"）。
                - 星期: str — 英文星期名称（如"Monday"）。
        """
        weekday_mapping = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6}
        df["春节标签"] = ""
        anchor_weeks = []

        for idx in df[df["农历"].str.contains("正月初一")].index:
            weekday = df.at[idx, "星期"]
            week_start = idx - weekday_mapping[weekday]
            week_end = week_start + 6
            anchor_weeks.append((week_start, week_end))

        anchor_weeks.sort()

        for i in range(len(df)):
            min_distance = float('inf')
            label = ""
            for start, end in anchor_weeks:
                if start <= i <= end:
                    min_distance = 0
                    label = "春节当周"
                    break
                elif i < start:
                    distance = (start - i - 1) // 7 + 1
                    if distance < min_distance:
                        min_distance = distance
                        label = f"节前{distance}周"
                elif i > end:
                    distance = (i - end - 1) // 7 + 1
                    if distance < min_distance:
                        min_distance = distance
                        label = f"节后{distance}周"
            df.at[i, "春节标签"] = label
        return df

    @staticmethod
    def generate_lunar_table_with_tags(start_year, end_year):
        """
        功能说明：
            生成指定年份范围内的农历日期对照表，并添加春节标签。
            遍历每一天，使用 ZhDate 库将阳历转换为农历日期，
            然后调用 add_global_spring_festival_tags 生成春节标签。

        params:
            start_year: int — 起始年份。
            end_year: int — 结束年份（包含）。
        """
        all_dates = []
        for year in range(start_year, end_year + 1):
            dt = datetime(year, 1, 1)
            while dt.year == year:
                try:
                    lunar = ZhDate.from_datetime(dt)
                    lunar_str = lunar.chinese()
                except:
                    lunar_str = "转换失败"
                all_dates.append({
                    "阳历": dt.strftime("%Y-%m-%d"),
                    "农历": lunar_str,
                    "星期": dt.strftime("%A"),
                    "年份": year
                })
                dt += timedelta(days=1)
        df_lunar = pd.DataFrame(all_dates)
        df_lunar = MacroPlugin.add_global_spring_festival_tags(df_lunar)
        return df_lunar

    @staticmethod
    def macro_create_festival_pivot(context, config):
        """
        功能说明：
            生成基于春节标签的透视表。
            从工作表中读取日期和数值数据，结合农历日期计算春节相对周数，
            按"春节标签"为行、年份为列生成透视表，用于分析春节前后数据变化趋势。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            config: dict — 春节透视表配置，包含：
                - base_col: str | int — 日期列（列字母或列号）。
                - value_col: str | int — 数值列（列字母或列号）。
                - target_start_col: str | int — 透视表写入的起始列。
                - target_start_row: int — 透视表写入的起始行。
                - recent_years: int (默认 7) — 包含的最近年份数量。
                - start_row: int (默认 10) — 数据读取的起始行号。
        """
        ws = context['ws']
        base_col = config['base_col']
        value_col = config['value_col']
        target_start_col = config['target_start_col']
        target_start_row = config['target_start_row']
        recent_years = config.get('recent_years', 7)
        
        # 1. 提取当前工作表 base_col 的最小最大日期，确定范围
        date_col_num = column_index_from_string(base_col) if isinstance(base_col, str) else base_col
        value_col_num = column_index_from_string(value_col) if isinstance(value_col, str) else value_col
        
        row = config.get('start_row', 10)
        dates = []
        date_to_value = {}
        
        while True:
            date_val = ws.cell(row=row, column=date_col_num).value
            val = ws.cell(row=row, column=value_col_num).value
            if date_val is None:
                break
            try:
                date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
                dates.append(pd.to_datetime(date_val))
                date_to_value[date_str] = val
            except:
                pass
            row += 1
            
        if not dates:
            print(f"[{ws.title}] 未找到数据以生成春节透视表")
            return
            
        max_year = max(dates).year
        start_year = max_year - 8
        
        df_lunar = MacroPlugin.generate_lunar_table_with_tags(start_year, max_year)
        
        tag_range = [f"节前{i}周" for i in range(6, 0, -1)] + ["春节当周"] + [f"节后{i}周" for i in range(1, 13)]
        df = df_lunar[df_lunar["春节标签"].isin(tag_range)].copy()
        df_result = df[["阳历", "春节标签"]].copy()
        df_result.columns = ["日期", "春节标签"]
        
        df_result["数值"] = df_result["日期"].map(date_to_value)
        df_result.loc[df_result["数值"] == 0, "数值"] = None
        df_result = df_result[df_result["数值"].notnull()].reset_index(drop=True)
        
        pivot_df = MacroPlugin.build_festival_pivot_table(df_result, recent_years=recent_years)
        
        # 写入工作表
        MacroPlugin.write_df_to_ws(ws, pivot_df, target_start_col, target_start_row, header=True)
        print(f"[{ws.title}] 生成春节标签透视表到 {target_start_col}")

    @staticmethod
    def build_festival_pivot_table(df_table, recent_years=7, value_col='数值'):
        """
        功能说明：
            构建春节标签透视表的核心逻辑。
            将带有春节标签的日期数据按标签和年份进行透视，
            处理跨年数据（10月之后的标签取前一年同标签数据）。

        params:
            df_table: pd.DataFrame — 包含日期、春节标签、数值的 DataFrame。
            recent_years: int (默认 7) — 包含的最近年份数量。
            value_col: str (默认 '数值') — 数值列名称。
        """
        df_table = df_table.copy()
        df_table['春节标签'] = df_table['春节标签'].str.strip()
        df_table['年份'] = pd.to_datetime(df_table['日期']).dt.year.astype(int)
        df_table['月份'] = pd.to_datetime(df_table['日期']).dt.month.astype(int)

        tag_list = [f"节前{i}周" for i in range(6, 0, -1)] + ["春节当周"] + [f"节后{i}周" for i in range(1, 13)]
        
        all_years = sorted(df_table['年份'].dropna().unique())
        years = all_years[-recent_years:]

        result = pd.DataFrame({'春节标签': tag_list})
        for y in years:
            result[str(y)] = None
        result = result.set_index('春节标签')

        for tag in tag_list:
            for year in years:
                rows = df_table[(df_table['春节标签'] == tag) & (df_table['年份'] == year)]
                if not rows.empty:
                    row = rows.iloc[0]
                    if row[value_col] is not None:
                        if row['月份'] > 9:
                            prev_rows = df_table[(df_table['春节标签'] == tag) & (df_table['年份'] == year-1)]
                            if not prev_rows.empty:
                                prev_row = prev_rows.sort_values('日期', ascending=False).iloc[0]
                                result.at[tag, str(year)] = prev_row[value_col]
                            else:
                                result.at[tag, str(year)] = None
                        else:
                            result.at[tag, str(year)] = row[value_col]
                    else:
                        if row['月份'] > 9:
                            prev_rows = df_table[(df_table['春节标签'] == tag) & (df_table['年份'] == year-1)]
                            if not prev_rows.empty:
                                prev_row = prev_rows.sort_values('日期', ascending=False).iloc[0]
                                result.at[tag, str(year)] = prev_row[value_col]
                            else:
                                result.at[tag, str(year)] = None
                        else:
                            result.at[tag, str(year)] = None
                else:
                    other_rows = df_table[(df_table['春节标签'] == tag)]
                    if not other_rows.empty:
                        prev_year_rows = other_rows[other_rows['年份'] == year-1]
                        typical_month = prev_year_rows.sort_values('日期', ascending=False).iloc[0]['月份'] if not prev_year_rows.empty else 1
                        if typical_month > 9:
                            prev_rows = df_table[(df_table['春节标签'] == tag) & (df_table['年份'] == year-1)]
                            if not prev_rows.empty:
                                prev_row = prev_rows.sort_values('日期', ascending=False).iloc[0]
                                result.at[tag, str(year)] = prev_row[value_col]
                            else:
                                result.at[tag, str(year)] = None
                        else:
                            result.at[tag, str(year)] = None
                    else:
                        result.at[tag, str(year)] = None

        return result.reset_index()

    # ==================================
    # 除夕相关逻辑
    # ==================================
    @staticmethod
    def generate_chuxi_relative_table(start_year, end_year):
        """
        功能说明：
            生成指定年份范围内的除夕相对天数对照表。
            计算每一天相对于当年除夕（农历腊月三十或廿九）的天数差，
            生成"t-N"（除夕前N天）、"t"（除夕当天）、"t+N"（除夕后N天）等标签。

        params:
            start_year: int — 起始年份。
            end_year: int — 结束年份（包含）。
        """
        all_rows = []
        for year in range(start_year, end_year + 1):
            try:
                chuxi = ZhDate(year-1, 12, 30).to_datetime()
            except Exception:
                chuxi = ZhDate(year-1, 12, 29).to_datetime()
            dt = datetime(year, 1, 1)
            while dt.year == year:
                try:
                    lunar_str = ZhDate.from_datetime(dt).chinese()
                except:
                    lunar_str = "转换失败"
                delta = (dt - chuxi).days
                if delta == 0: rel = "t"
                elif delta < 0: rel = f"t{delta}"
                else: rel = f"t+{delta}"
                all_rows.append({
                    "阳历": dt.strftime("%Y-%m-%d"),
                    "农历": lunar_str,
                    "除夕相对天数": rel
                })
                dt += timedelta(days=1)
        return pd.DataFrame(all_rows)

    @staticmethod
    def macro_create_chuxi_pivot(context, config):
        """
        功能说明：
            生成基于除夕相对天数的透视表。
            从工作表中读取日期和数值数据，结合农历日期计算除夕相对天数，
            按"除夕相对天数"为行、年份为列生成透视表。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            config: dict — 除夕透视表配置，包含：
                - date_col: str | int — 日期列（列字母或列号）。
                - value_col: str | int — 数值列（列字母或列号）。
                - target_start_col: str | int — 透视表写入的起始列。
                - target_start_row: int — 透视表写入的起始行。
                - start_row: int (默认 11) — 数据读取的起始行号。
        """
        ws = context['ws']
        date_col = config['date_col']
        value_col = config['value_col']
        target_start_col = config['target_start_col']
        target_start_row = config['target_start_row']
        
        date_col_num = column_index_from_string(date_col) if isinstance(date_col, str) else date_col
        value_col_num = column_index_from_string(value_col) if isinstance(value_col, str) else value_col
        
        row = config.get('start_row', 11)
        dates = []
        date_to_value = {}
        
        while True:
            date_val = ws.cell(row=row, column=date_col_num).value
            val = ws.cell(row=row, column=value_col_num).value
            if date_val is None:
                break
            try:
                date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
                dates.append(pd.to_datetime(date_val))
                date_to_value[date_str] = val
            except:
                pass
            row += 1
            
        if not dates:
            print(f"[{ws.title}] 未找到数据以生成除夕透视表")
            return
            
        max_year = max(dates).year
        start_year = max_year - 8
        
        df = MacroPlugin.generate_chuxi_relative_table(start_year, max_year)
        df_chuxi = df.copy()
        df_chuxi["数值"] = df_chuxi["阳历"].map(date_to_value)
        
        # build_chuxi_relative_pivot_table
        chuxi_tags = [f"t-{i}" for i in range(30, 0, -1)] + ["t"] + [f"t+{i}" for i in range(1, 61)]
        years = list(range(max_year-6, max_year+1))
        
        result = pd.DataFrame(index=chuxi_tags, columns=years)
        df_chuxi['年份'] = pd.to_datetime(df_chuxi['阳历']).dt.year
        
        for tag in chuxi_tags:
            for year in years:
                r = df_chuxi[(df_chuxi['除夕相对天数'] == tag) & (df_chuxi['年份'] == year)]
                if not r.empty:
                    result.at[tag, year] = r.iloc[0]["数值"]
                else:
                    result.at[tag, year] = None
        result = result.reset_index()
        
        MacroPlugin.write_df_to_ws(ws, result, target_start_col, target_start_row, header=True)
        print(f"[{ws.title}] 生成除夕相对天数透视表到 {target_start_col}")

    @staticmethod
    def macro_create_weekly_pivot(context, config):
        """
        功能说明：
            生成按 ISO 周统计的数据透视表。
            从工作表中读取日期和数值数据，提取 ISO 年份和周数，
            按"周数"为行、年份为列生成透视表。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            config: dict — 周度透视表配置，包含：
                - date_col: str | int — 日期列（列字母或列号）。
                - value_col: str | int — 数值列（列字母或列号）。
                - target_start_col: str | int — 透视表写入的起始列。
                - target_start_row: int — 透视表写入的起始行。
                - start_year: int (默认 2015) — 数据起始年份。
                - end_year: int (默认 2024) — 数据结束年份。
                - start_row: int (默认 2) — 数据读取的起始行号。
        """
        ws = context['ws']
        date_col = config['date_col']
        value_col = config['value_col']
        target_start_col = config['target_start_col']
        target_start_row = config['target_start_row']
        start_year = config.get('start_year', 2015)
        end_year = config.get('end_year', 2024)
        
        date_col_num = column_index_from_string(date_col) if isinstance(date_col, str) else date_col
        value_col_num = column_index_from_string(value_col) if isinstance(value_col, str) else value_col
        
        row = config.get('start_row', 2)
        data = []
        
        while True:
            date_val = ws.cell(row=row, column=date_col_num).value
            value_val = ws.cell(row=row, column=value_col_num).value
            if date_val is None:
                break
            try:
                date_dt = pd.to_datetime(date_val)
                iso_year, iso_week, _ = date_dt.isocalendar()
                if start_year <= iso_year <= end_year:
                    data.append({'year': iso_year, 'week': iso_week, 'value': value_val})
            except:
                pass
            row += 1
            
        df = pd.DataFrame(data)
        weeks = list(range(1, 54))
        years = list(range(start_year, end_year + 1))
        result = pd.DataFrame(index=weeks, columns=years)
        
        if not df.empty:
            for year in years:
                for week in weeks:
                    vals = df[(df['year'] == year) & (df['week'] == week)]['value']
                    result.at[week, year] = vals.iloc[0] if not vals.empty else None
                    
        result.index.name = 'week'
        result.insert(0, 'index', range(1, 54))
        
        MacroPlugin.write_df_to_ws(ws, result, target_start_col, target_start_row, header=True)
        print(f"[{ws.title}] 生成按周数据透视表到 {target_start_col}")

    @staticmethod
    def macro_create_yearly_date_scaffold(context, config):
        """
        功能说明：
            生成按年分列的日期脚手架（日期框架）。
            为指定年份范围内的每一天生成日期行，每年占用两列（一列日期，一列预留数值），
            用于后续填充数据。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            config: dict — 日期脚手架配置，包含：
                - start_year: int — 起始年份。
                - end_year: int — 结束年份（包含）。
                - pivot_start_col: str | int — 透视表写入的起始列。
                - pivot_start_row: int — 透视表写入的起始行。
                - row: int — 年份标题所在行号。
        """
        ws = context['ws']
        start_year = config['start_year']
        end_year = config['end_year']
        pivot_start_col = config['pivot_start_col']
        pivot_start_row = config['pivot_start_row']
        row = config['row']
        
        pivot_col_num = column_index_from_string(pivot_start_col) if isinstance(pivot_start_col, str) else pivot_start_col
        
        for i, year in enumerate(range(start_year, end_year + 1)):
            year_dates = pd.date_range(start=f"{year}-01-01", end=f"{year}-12-31", freq='D')
            year_col1 = pivot_col_num + i * 2
            
            ws.cell(row=row, column=year_col1, value=f"{year}")
            
            for j, date_val in enumerate(year_dates):
                row_idx = pivot_start_row + j
                cell = ws.cell(row=row_idx, column=year_col1, value=pd.to_datetime(date_val))
                cell.number_format = 'yyyy-mm-dd'
                
        print(f"[{ws.title}] 生成按年日期脚手架到 {pivot_start_col} ({start_year}-{end_year})")

    @staticmethod
    def write_df_to_ws(ws, df, start_col, start_row, header=True):
        """
        功能说明：
            将 DataFrame 写入工作表的指定位置。
            支持写入表头和数据行，自动将 0 值替换为 None（空单元格）。

        params:
            ws: openpyxl.Worksheet — 目标工作表对象。
            df: pd.DataFrame — 要写入的 DataFrame。
            start_col: str | int — 写入的起始列（列字母或列号）。
            start_row: int — 写入的起始行号。
            header: bool (默认 True) — 是否写入表头行。
        """
        col_num = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
        df = df.replace(0, None)
        
        if header:
            for j, col_name in enumerate(df.columns):
                ws.cell(row=start_row, column=col_num + j, value=col_name)
            data_start_row = start_row + 1
        else:
            data_start_row = start_row

        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                ws.cell(row=data_start_row + i, column=col_num + j, value=value)