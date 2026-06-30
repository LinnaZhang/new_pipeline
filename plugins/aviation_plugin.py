import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

from core_engine.data_processor import DataProcessor

class AviationPlugin:
    @staticmethod
    def _find_section_rows(ws):
        """
        扫描工作表的 A 列，定位"季度"、"YTD"、"全年"等关键区域的行号。

        Args:
            ws: openpyxl.Worksheet — 待扫描的工作表对象。

        Returns:
            dict:
                - quarter_start_row: int | None — "季度"标签所在行号。
                - ytd_row: int | None — "YTD"标签所在行号。
                - full_year_row: int | None — "全年"标签所在行号。
                - last_label_row: int — A 列最后一个非空文本行号。

        Usage:
            仅供 AviationPlugin 内部方法调用。
        """
        quarter_start_row = None
        ytd_row = None
        full_year_row = None
        last_label_row = 0

        for row_idx in range(1, ws.max_row + 1):
            raw_value = ws.cell(row=row_idx, column=1).value
            if raw_value is None:
                continue

            value = str(raw_value).strip()
            if not value:
                continue

            last_label_row = row_idx
            if value == "季度":
                quarter_start_row = row_idx
            elif "YTD" in value.upper():
                ytd_row = row_idx
            elif "全年" in value:
                full_year_row = row_idx

        return {
            "quarter_start_row": quarter_start_row,
            "ytd_row": ytd_row,
            "full_year_row": full_year_row,
            "last_label_row": last_label_row,
        }

    @staticmethod
    def _clear_columns_below_row(ws, cols_to_clear, start_row):
        """
        清空指定行及之后所有行中指定列的内容（将单元格值设为 None）。
        常用于清除旧的公式或数据，防止写入新内容时残留脏数据。

        Args:
            ws: openpyxl.Worksheet — 目标工作表对象。
            cols_to_clear: list[str] — 需要清空的列字母列表，如 ["B", "D", "H"]。
            start_row: int — 起始行号（包含），从此行开始向下清空。

        Returns:
            None: 原地修改传入的 ws 对象。
        """
        if not start_row or start_row > ws.max_row:
            return

        for col in cols_to_clear:
            col_num = column_letter_to_number(col)
            for row_idx in range(start_row, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_num).value = None

    @staticmethod
    def _get_last_formula_row(ws):
        """
        获取工作表中最后一个有效标签行号，同时返回区域行信息。

        封装 `_find_section_rows`，仅提取 last_label_row 和完整的区域映射，
        供调用方判断数据边界和跳行。

        Args:
            ws: openpyxl.Worksheet — 目标工作表对象。

        Returns:
            tuple:
                - int: 最后一个非空标签行号（last_label_row）。
                - dict: `_find_section_rows` 的完整返回值。

        Usage:
            仅供 AviationPlugin 内部方法调用。
        """
        section_rows = AviationPlugin._find_section_rows(ws)
        return section_rows["last_label_row"], section_rows
    
    # ==========================================================================
    # 基础数据写入方法
    # ==========================================================================

    @staticmethod
    def aviation_write_airline_sheet(context, params):
        """
        处理航空公司基础指标数据，写入到目标工作表中，并自动生成季度统计和 YTD 统计。

        核心流程：
        1. 从 data_reader 缓存中读取多个指标数据。
        2. 以第一个指标的日期为基准，自动补全缺失月份（月末日期），生成统一日期主表。
        3. 在 Excel 中写入日期列和各指标数据列，支持单位转换。
        4. 生成"季度"区域的季度汇总统计。
        5. 生成 YTD 区域的全年累计统计。
        6. 对国泰航空数据额外处理缺失值（列H = B - E，列Q = K - N）。
        7. 对指定列（AL, AO, AR, AU）设置两位小数格式。

        Args:
            context: dict — Pipeline 上下文，包含：
                - ws: openpyxl.Worksheet — 目标工作表。
                - data_reader: DataReader — 指标数据读取器。
                - sheet_config: dict — 工作表级配置，包含：
                    - source_sheet: str — 数据源工作表名。
                    - indicators: dict[str, str] — 指标名到列字母的映射。
                    - unit_conversion: dict[str, float] | None — 列字母到倍率的单位转换映射。
            params: dict — 从 YAML action 配置中解析的参数，包含：
                - start_row: int (默认 4) — 数据写入的起始行号。
                - start_date: str (默认 '2014-01-01') — 数据最早日期。
                - date_format: str (默认 'yyyy-mm') — 日期列的数字格式。

        Returns:
            None: 直接在传入的 ws 对象上原地写入数据。

        Usage:
            由 Pipeline 引擎根据 YAML 配置调用，通常作为航空公司基础数据表的第一步操作。
        """
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']

        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 4)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm')

        # 获取单位转换配置，如果为 None 则设为空字典
        unit_conversion = sheet_config.get('unit_conversion') or {}

        start_date_ts = pd.to_datetime(start_date)

        def convert_value(column_letter, value, conversion_map):
            """根据列和配置的倍率转换数值"""
            if value is None or not isinstance(value, (int, float)):
                return value
            # 确保 conversion_map 是字典类型
            if conversion_map and column_letter in conversion_map:
                return value * conversion_map[column_letter]
            return value

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 构造统一日期主表 base_df
        #    逻辑：
        #    - 以第一个指标为主指标
        #    - 如果第一个指标最早日期晚于 start_date 所在月份，则自动补全中间缺失月份
        #    - 补全日期使用每个月的月末日期，例如 2014-01-31、2014-02-28
        first_indicator = list(indicator_col_map.keys())[0]
        first_col = indicator_col_map[first_indicator]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        first_real_date = pd.to_datetime(first_df['日期'].iloc[0])

        # start_date 对应月份的月末，例如 2014-01-01 -> 2014-01-31
        start_month_end = start_date_ts + pd.offsets.MonthEnd(0)

        # 第一个真实日期所在月份的上一个月月末
        first_real_month_end = first_real_date + pd.offsets.MonthEnd(0)
        fill_end = first_real_month_end - pd.offsets.MonthEnd(1)

        # 如果第一个真实日期晚于起始月份，则补全 start_date 到真实日期前一个月
        if fill_end >= start_month_end:
            fill_dates = pd.date_range(
                start=start_month_end,
                end=fill_end,
                freq='ME'
            )

            fill_df = pd.DataFrame({'日期': fill_dates})

            # 补齐 first_df 的其他列，避免 concat 后列不一致
            for col in first_df.columns:
                if col not in fill_df.columns:
                    fill_df[col] = None

            # 保持列顺序和 first_df 一致
            fill_df = fill_df[first_df.columns]

            base_df = pd.concat([fill_df, first_df], ignore_index=True)
        else:
            base_df = first_df.copy()

        base_df = base_df.sort_values('日期', ascending=True).reset_index(drop=True)

        # 3. 写入统一日期列和第一个指标数据（应用单位转换）
        first_value_col = first_df.columns[-1]
        first_value_map = dict(zip(first_df['日期'], first_df[first_value_col]))

        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            # A列写日期
            date_cell = ws.cell(row=current_row, column=1, value=current_date)
            date_cell.number_format = date_format

            # 第一个指标列写值，应用单位转换
            value = first_value_map.get(current_date, None)
            converted_value = convert_value(first_col, value, unit_conversion)
            ws.cell(
                row=current_row,
                column=column_letter_to_number(first_col),
                value=converted_value
            )

        # 4. 写入其他指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items())[1:]:
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]

            value_map = dict(zip(df['日期'], df[value_col]))

            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                converted_value = convert_value(col, value, unit_conversion)
                ws.cell(
                    row=start_row + i,
                    column=column_letter_to_number(col),
                    value=converted_value
                )
        
        # 4.1 处理国泰航空缺失值
        # 判断当前是否为国泰航空数据
        if source_sheet == "国泰航空":
            print("    - 处理国泰航空缺失值")
            for i, date in enumerate(base_date_list):
                current_row = start_row + i
                
                # 获取列B (第2列) 和列E (第5列)
                col_b_value = ws.cell(row=current_row, column=2).value
                col_e_value = ws.cell(row=current_row, column=5).value
                # 若列B和列E存在值，则列H = B - E
                if col_b_value is not None and col_e_value is not None:
                    if col_b_value is not None and col_e_value is not None:
                        col_h_value = col_b_value - col_e_value
                        ws.cell(row=current_row, column=8, value=col_h_value)  # 列H是第8列
                
                # 获取列K (第11列) 和列N (第14列)
                col_k_value = ws.cell(row=current_row, column=11).value
                col_n_value = ws.cell(row=current_row, column=14).value
                
                # 若列K和列N存在值，则列Q = K - N
                if col_k_value is not None and col_n_value is not None:
                    if col_k_value is not None and col_n_value is not None:
                        col_q_value = col_k_value - col_n_value
                        ws.cell(row=current_row, column=17, value=col_q_value)  # 列Q是第17列

        # 5. 添加季度统计标题
        empty_row = start_row + len(base_df)
        ws.cell(row=empty_row, column=1, value=None)

        title_row = empty_row + 1
        ws.cell(row=title_row, column=1, value="季度")
        ws.cell(row=title_row, column=1).font = openpyxl.styles.Font(bold=True)

        # 6. 构造季度统计数据源
        #    注意：这里必须基于 base_df，而不是 first_df
        #    否则补全出来的日期不会进入季度统计，导致季度结果错误
        valid_dates = pd.to_datetime(base_df['日期'], errors='coerce')
        df_2018 = base_df[valid_dates.notna() & (valid_dates.dt.year >= 2018)].copy()
        df_2018.reset_index(drop=True, inplace=True)

        df_all_cols = pd.DataFrame({'日期': df_2018['日期']})
        quarter_date_list = df_2018['日期'].tolist()

        for indicator, col in indicator_col_map.items():
            indicator_df = indicator_dfs[indicator].copy()

            if indicator_df.empty:
                df_all_cols[col] = [None for _ in quarter_date_list]
                continue

            indicator_df['日期'] = pd.to_datetime(indicator_df['日期'], errors='coerce')
            indicator_df = indicator_df[indicator_df['日期'].notna()].copy()

            value_col = indicator_df.columns[-1]
            value_map = dict(zip(indicator_df['日期'], indicator_df[value_col]))

            # 季度统计时也应用单位转换
            df_all_cols[col] = [
                convert_value(col, value_map.get(date, None), unit_conversion)
                for date in quarter_date_list
            ]

        processor = DataProcessor(reader)

        quarter_title_row, quarter_results = processor.add_quarterly_stats(
            ws=ws,
            start_row=start_row,
            data_df=df_all_cols,
            value_cols=list(indicator_col_map.values())
        )

        current_row = title_row + 1
        for quarter_data in quarter_results:
            if quarter_data['quarter_label']:
                ws.cell(row=current_row, column=1, value=quarter_data['quarter_label'])

                for col, value in quarter_data['data'].items():
                    cell = ws.cell(
                        row=current_row,
                        column=column_letter_to_number(col),
                        value=value
                    )

                    if value is not None:
                        original_col = ws.cell(
                            row=start_row,
                            column=column_letter_to_number(col)
                        )
                        cell.number_format = original_col.number_format

            current_row += 1

        last_quarter_row = current_row - 1
        while last_quarter_row > title_row:
            cell_value = ws.cell(row=last_quarter_row, column=1).value
            if cell_value and 'Q' in str(cell_value):
                break
            last_quarter_row -= 1

        last_ytd_row = processor.add_ytd_stats(
            ws=ws,
            last_quarter_row=last_quarter_row,
            data_df=df_all_cols,
            value_cols=list(indicator_col_map.values())
        )

        # 7. 指定列保留两位小数
        target_decimal_cols = ['AL', 'AO', 'AR', 'AU']
        for col in target_decimal_cols:
            if col in indicator_col_map.values():
                col_num = column_letter_to_number(col)

                for row in range(start_row, last_ytd_row + 1):
                    cell = ws.cell(row=row, column=col_num)

                    if cell.value is not None:
                        cell.number_format = '0.00'

        print(f"    - 完成航空基础数据写入与季度统计")
        
    @staticmethod
    def aviation_apply_yoy_formulas(context, params):
        """
        功能说明：
            计算并写入各指标的同比上年增长率公式（YoY）。
            根据日期标签类型（年/季度/月份），计算当期值与上年同期值的增长率公式：
        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - target_cols: list[str] — 需要计算同比的目标列字母列表。
                - start_row: int (默认 16) — 数据起始行号。        """
        ws = context['ws']
        target_cols = params['target_cols']
        start_row = params.get('start_row', 16)

        last_formula_row, section_rows = AviationPlugin._get_last_formula_row(ws)
        full_year_row = section_rows["full_year_row"]
        AviationPlugin._clear_columns_below_row(ws, target_cols, last_formula_row + 1)

        data_2017_rows = {}
        for row_idx in range(start_row, last_formula_row + 1):
            raw_value = ws.cell(row=row_idx, column=1).value
            cell_value = str(raw_value).strip() if raw_value is not None else ""
            if cell_value.startswith("2017-"):
                try:
                    month = int(cell_value.split("-")[1])
                    data_2017_rows[month] = row_idx
                except (ValueError, IndexError):
                    continue

        for col in target_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
            
            for row in range(start_row, last_formula_row + 1):
                formula = ""
                current_cell = ws.cell(row=row, column=column_letter_to_number(source_col))
                raw_date_value = ws.cell(row=row, column=1).value
                date_value = str(raw_date_value).strip() if raw_date_value is not None else ""
                if not date_value or pd.isna(current_cell.value):
                    continue
                
                if "年" in date_value:
                    if full_year_row is not None and row > full_year_row:
                        prev_row = row - 1
                        if prev_row >= 1:
                            formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                    else:
                        try:
                            year = int(date_value.replace("年", ""))
                            if year <= 2018:
                                formula = ""
                            else:
                                prev_row = row - 2
                                if prev_row >= 1:
                                    formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                        except (ValueError, IndexError):
                            prev_row = row - 2
                            if prev_row >= 1:
                                formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                elif "18" in date_value and "Q" in date_value:
                    try:
                        quarter = int(date_value[date_value.find("Q") + 1])
                        q_month_rows = []
                        for month in range((quarter - 1) * 3 + 1, (quarter - 1) * 3 + 4):
                            if month in data_2017_rows:
                                q_month_rows.append(data_2017_rows[month])

                        if len(q_month_rows) == 3:
                            sum_formula = f"(SUM({source_col}{q_month_rows[0]}:{source_col}{q_month_rows[2]}))"
                            formula = f"={source_col}{row}/{sum_formula}-1"
                    except (ValueError, IndexError):
                        formula = ""
                elif "Q" in date_value:
                    prev_row = row - 5
                    if prev_row >= 1:
                        formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                else:
                    prev_row = row - 12
                    if prev_row >= 1:
                        formula = f"={source_col}{row}/{source_col}{prev_row}-1"
                prev_cell = ws.cell(row=prev_row, column=column_letter_to_number(source_col))

                if pd.isna(prev_cell.value):
                    formula = ""

                if formula:
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = formula
                    cell.number_format = "0.00%"
        print(f"    - 完成航空同比上年计算: {target_cols}")

    @staticmethod
    def aviation_apply_yoy_diff_formulas(context, params):
        """
        功能说明：
            计算并写入各指标的同比上年差值公式（YoY Diff）。
            计算当期值与上年同期值的差值.

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - target_cols: list[str] — 需要计算同比差值的列字母列表。
                - start_row: int (默认 16) — 数据起始行号。
        """
        ws = context['ws']
        target_cols = params['target_cols']
        start_row = params.get('start_row', 16)
        last_formula_row, _ = AviationPlugin._get_last_formula_row(ws)
        AviationPlugin._clear_columns_below_row(ws, target_cols, last_formula_row + 1)
            
        for col in target_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 1)
            for row in range(start_row, last_formula_row + 1):
                formula = ""
                raw_date_value = ws.cell(row=row, column=1).value
                date_value = str(raw_date_value).strip() if raw_date_value is not None else ""

                current_cell = ws.cell(row=row, column=column_letter_to_number(source_col))
                if not date_value or pd.isna(current_cell.value):
                    continue
                
                if not date_value:
                    continue
                if "年" in date_value:
                    if "2018年" in date_value:
                        continue
                    prev_row = row - 2
                    if prev_row >= 1:
                        formula = f"={source_col}{row}-{source_col}{prev_row}"
                elif "Q" in date_value:
                    if "18" in date_value:
                        formula = f"={source_col}{row}-AVERAGE({source_col}40:{source_col}42)"
                    prev_row = row - 5
                    if prev_row >= 1:
                        formula = f"={source_col}{row}-{source_col}{prev_row}"
                elif isinstance(date_value, datetime):
                    prev_row = row - 12
                    if prev_row >= 1:
                        formula = f"={source_col}{row}-{source_col}{prev_row}"
                else:
                    try:
                        datetime.strptime(str(date_value), "%Y-%m-%d %H:%M:%S")

                        prev_row = row - 12
                        if prev_row >= 1:
                            formula = f"={source_col}{row}-{source_col}{prev_row}"
                        else:
                            formula = ""
                    except ValueError:
                        formula = ""
                        continue

                prev_cell = ws.cell(row=prev_row, column=column_letter_to_number(source_col))                
                if pd.isna(prev_cell.value):
                    formula = ""

                if formula:
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = formula
                    cell.number_format = "0.00"
        print(f"    - 完成航空同比差值计算: {target_cols}")

    @staticmethod
    def aviation_apply_yoy19_formulas(context, params):
        """
        功能说明：
            计算并写入各指标相对于2019年同期的增长率公式（YoY vs 2019）。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - target_cols: list[str] — 需要计算同比2019的目标列字母列表。
                - start_row: int (默认 112) — 数据起始行号。
        """
        ws = context['ws']
        target_cols = params['target_cols']
        start_row = params.get('start_row', 112)
        last_formula_row, section_rows = AviationPlugin._get_last_formula_row(ws)
        full_year_row = section_rows["full_year_row"]
        AviationPlugin._clear_columns_below_row(ws, target_cols, last_formula_row + 1)

        for col in target_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
            
            # 找到2019年相关行
            row_2019_base = 64
            row_2019 = None
            row_2019_after_full_year = None
            row_2019Q1 = row_2019Q2 = row_2019Q3 = row_2019Q4 = None
            full_year_row_19 = None
            
            for r in range(1, ws.max_row + 1):
                val = str(ws.cell(row=r, column=1).value)
                if val == "2019年":
                    if full_year_row_19 and r > full_year_row_19:
                        row_2019_after_full_year = r
                    else:
                        row_2019 = r
                elif val == "19Q1": row_2019Q1 = r
                elif val == "19Q2": row_2019Q2 = r
                elif val == "19Q3": row_2019Q3 = r
                elif val == "19Q4": row_2019Q4 = r
                elif "全年" in val: full_year_row_19 = r

            for row in range(start_row, last_formula_row + 1):
                formula = ""
                raw_date_value = ws.cell(row=row, column=1).value
                date_value = str(raw_date_value).strip() if raw_date_value is not None else ""
                if not date_value:
                    continue
                
                if "Q" in date_value and "19" not in date_value:
                    try:
                        parts = date_value.split("Q")
                        if len(parts) == 2:
                            year_part = parts[0]
                            quarter = parts[1]
                            if year_part.isdigit():
                                current_year = int(year_part)
                                if current_year >= 23:
                                    q_row_map = {"1": row_2019Q1, "2": row_2019Q2, "3": row_2019Q3, "4": row_2019Q4}
                                    row_19_q = q_row_map.get(quarter)
                                    if row_19_q:
                                        formula = f"={source_col}{row}/{source_col}{row_19_q}-1"
                    except IndexError:
                        continue
                elif "年" in date_value:
                    year_str = "".join(filter(str.isdigit, date_value))
                    if year_str and int(year_str) >= 2023:
                        if full_year_row and row > full_year_row and row_2019_after_full_year:
                            formula = f"={source_col}{row}/{source_col}{row_2019_after_full_year}-1"
                        elif row_2019:
                            formula = f"={source_col}{row}/{source_col}{row_2019}-1"
                elif "19" not in date_value:
                    month = None
                    try:
                        month = pd.to_datetime(date_value).month
                    except: pass
                    if month:
                        base_row = row_2019_base + month - 1
                        formula = f"={source_col}{row}/{source_col}{base_row}-1"

                if formula:
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = formula
                    cell.number_format = "0.00%"
        print(f"    - 完成航空同比19计算: {target_cols}")

    @staticmethod
    def aviation_apply_diff19_formulas(context, params):
        """
        功能说明：
            计算并写入各指标相对于2019年同期的差值公式（Diff vs 2019）。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - target_cols: list[str] — 需要计算差值型同比2019的目标列字母列表。
                - start_row: int (默认 112) — 数据起始行号。
        """        
        ws = context['ws']
        target_cols = params['target_cols']
        start_row = params.get('start_row', 112)
        last_formula_row, section_rows = AviationPlugin._get_last_formula_row(ws)
        full_year_row = section_rows["full_year_row"]
        AviationPlugin._clear_columns_below_row(ws, target_cols, last_formula_row + 1)

        for col in target_cols:
            col_num = column_letter_to_number(col)
            source_col = column_number_to_letter(col_num - 2)
            
            row_2019_base = 64
            row_2019 = None
            row_2019_after_full_year = None
            row_2019Q1 = row_2019Q2 = row_2019Q3 = row_2019Q4 = None
            full_year_row_19 = None
            
            for r in range(1, ws.max_row + 1):
                val = str(ws.cell(row=r, column=1).value)
                if val == "2019年":
                    if full_year_row_19 and r > full_year_row_19:
                        row_2019_after_full_year = r
                    else:
                        row_2019 = r
                elif val == "19Q1": row_2019Q1 = r
                elif val == "19Q2": row_2019Q2 = r
                elif val == "19Q3": row_2019Q3 = r
                elif val == "19Q4": row_2019Q4 = r
                elif "全年" in val: full_year_row_19 = r

            for row in range(start_row, last_formula_row + 1):
                formula = ""
                raw_date_value = ws.cell(row=row, column=1).value
                date_value = str(raw_date_value).strip() if raw_date_value is not None else ""
                if not date_value:
                    continue
                
                if "Q" in date_value:
                    year = date_value[:2]
                    if year >= "23":
                        quarter = date_value[-1]
                        q_row_map = {"1": row_2019Q1, "2": row_2019Q2, "3": row_2019Q3, "4": row_2019Q4}
                        row_19_q = q_row_map.get(quarter)
                        if row_19_q:
                            formula = f"={source_col}{row}-{source_col}{row_19_q}"
                elif "年" in date_value:
                    year_str = "".join(filter(str.isdigit, date_value))
                    if year_str and int(year_str) >= 2023:
                        if full_year_row and row > full_year_row and row_2019_after_full_year:
                            formula = f"={source_col}{row}-{source_col}{row_2019_after_full_year}"
                        elif row_2019:
                            formula = f"={source_col}{row}-{source_col}{row_2019}"
                elif "19" not in date_value:
                    month = None
                    try:
                        month = pd.to_datetime(date_value).month
                    except: pass
                    if month:
                        base_row = row_2019_base + month - 1
                        formula = f"={source_col}{row}-{source_col}{base_row}"

                if formula:
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = formula
                    cell.number_format = "0.00"
        print(f"    - 完成航空差值型同比19计算: {target_cols}")

    @staticmethod #清 2014–2017 年这些行里的同比/同比19列（让它们保持空白）。
    def aviation_clear_early_years_data(context, params):
        ws = context['ws']
        target_years = ["2014年", "2015年", "2016年", "2017年"]
        cols_to_clear = params.get('yoy_cols', []) + params.get('yoy19_cols', [])
        
        for row_idx in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_idx, column=1).value)
            if cell_value in target_years:
                for col in cols_to_clear:
                    ws.cell(row=row_idx, column=column_letter_to_number(col)).value = None
        print(f"    - 完成清除2014-2017年早年无用公式")

    @staticmethod #全年行以下的数据格式调整(保留一位小数)
    def aviation_adjust_format_after_full_year(context, params):
        ws = context['ws']
        cols = params.get('yoy_cols', []) + params.get('yoy19_cols', [])
        
        full_year_row = None
        for row_idx in range(1, ws.max_row + 1):
            if "全年" in str(ws.cell(row=row_idx, column=1).value):
                full_year_row = row_idx
                break
                
        if full_year_row:
            for col in cols:
                col_num = column_letter_to_number(col)
                for row in range(full_year_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_num)
                    if cell.value is not None:
                        cell.number_format = "0.0%"
        print(f"    - 完成全年行以下的数据格式调整(保留一位小数)")

    @staticmethod # 清除YTD中2018年的同比差值
    def aviation_clear_ytd_2018_diff_data(context, params):
        ws = context['ws']
        yoy_diff_cols = params.get('yoy_diff_cols', [])
        
        ytd_row = full_year_row = None
        for row_idx in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_idx, column=1).value)
            if "YTD" in val.upper(): ytd_row = row_idx
            elif "全年" in val: full_year_row = row_idx
            
        if ytd_row and full_year_row:
            for row_idx in range(ytd_row, full_year_row):
                if str(ws.cell(row=row_idx, column=1).value) == "2018年":
                    for col in yoy_diff_cols:
                        ws.cell(row=row_idx, column=column_letter_to_number(col)).value = None
        print(f"    - 完成清除YTD中2018年的同比差值")

    @staticmethod # 清除季度/YTD/全年数据的AX, AY, AZ列的残留（模板特殊列）
    def aviation_clear_ax_ay_az_columns(context, params):
        ws = context['ws']
        cols_to_clear = ["AX", "AY", "AZ"]
        
        quarter_start_row = ytd_row = full_year_row = None
        for row_idx in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_idx, column=1).value)
            if val == "季度": quarter_start_row = row_idx
            elif "YTD" in val.upper(): ytd_row = row_idx
            elif "全年" in val: full_year_row = row_idx
            
        for row_idx in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_idx, column=1).value)
            should_clear = False
            if quarter_start_row and ytd_row and quarter_start_row < row_idx < ytd_row and "Q" in val:
                should_clear = True
            elif ytd_row and full_year_row and ytd_row < row_idx < full_year_row and "年" in val:
                should_clear = True
            elif full_year_row and row_idx > full_year_row and "年" in val:
                should_clear = True
                
            if should_clear:
                for col in cols_to_clear:
                    cell = ws.cell(row=row_idx, column=column_letter_to_number(col))
                    # 无论原先有没有值，直接清除内容并把公式清空，防止出现 #NAME? 等错误
                    cell.value = None
        print(f"    - 完成清除季度/YTD/全年数据的AX, AY, AZ列")

    @staticmethod
    def get_latest_date_from_column_a(sheet_data):
        """
        从给定的 sheet 数据中筛选 A 列的日期，找到最新的日期。

        Args:
            sheet_data (pd.DataFrame): 包含 Excel 数据的 DataFrame，假设 A 列为 '日期'。

        Returns:
            tuple: (sheet_name, latest_date)，其中 latest_date 格式为 'YYYY-MM-DD'。
        """
        try:
            # 确保 A 列存在并转换为日期格式
            if 'A' in sheet_data.columns:
                sheet_data['A'] = pd.to_datetime(sheet_data['A'], errors='coerce')
                latest_date = sheet_data['A'].max()
                if pd.notna(latest_date):
                    return latest_date.strftime('%Y-%m-%d')
            return None
        except Exception as e:
            print(f"Error processing sheet data: {e}")
            return None
        
    @staticmethod
    def aviation_write_monthly_report_header_info(context, params):
        """
        处理月度汇总表，写入标题与参数列。
        """
        wb = context['wb']
        ws = context['ws']
        
        # 获取目标工作表名称列表
        target_sheets_name = params['target_sheets']
        # 获取标题列与行数，参数列与参数行数
        title_col = params['title_col']
        title_col_num = column_letter_to_number(title_col)
        title_row = params['title_row']
        items_col = params['items_col']
        items_col_num = column_letter_to_number(items_col)
        items_rows = params.get('items_rows', [])
        
        # 列出目标文件中的所有工作表名称
        sheet_names = wb.sheetnames
        
        # 存储每个工作表的最新日期
        latest_dates = {}
        latest_rows = {}

        for sheet_name in target_sheets_name:
            if sheet_name in sheet_names:
                # print(f"➡️ 正在处理工作表: {sheet_name}")
                ws = wb[sheet_name]
                
                # 遍历 A 列，获取最新日期
                dates = []
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value  # A 列
                    if isinstance(cell_value, datetime):  # 确保是日期类型
                        dates.append(cell_value)
                
                if dates:
                    latest_date = max(dates)  # 获取最新日期
                    # 获取最新日期的行数
                    latest_row = next((row_idx for row_idx in range(1, ws.max_row + 1) if ws.cell(row=row_idx, column=1).value == latest_date), None)
                    latest_rows[sheet_name] = latest_row
                    latest_dates[sheet_name] = latest_date
                    # print(f"    - 工作表 [{sheet_name}] 的最新日期: {latest_date.strftime('%Y-%m-%d')}")
                else:
                    print(f"    - 工作表 [{sheet_name}] 的 A 列没有有效日期数据")
            else:
                print(f"⚠️ 跳过不存在的工作表: {sheet_name}")
        
        # 检查所有工作表的最新日期是否一致
        if latest_dates:
            ws = context['ws']
            unified_date = min(latest_dates.values())
            unified_row = min(latest_rows.values())

            ws.cell(row=1, column=column_letter_to_number('C'), value=unified_row)
            print(f"✅ 所有工作表的最新日期一致: {unified_date.strftime('%Y-%m-%d')}")
            # 写入标题单元格
            title = f"【申万交运】{unified_date.strftime('%Y年%m月')}航空公司数据汇总"
            ws.cell(row=title_row, column=title_col_num, value=title)
            print(f"✅ 标题已写入: {title}")

            # 写入参数列
            # 对于 items_col_num 列，对于items_rows中的item_row，写入“当期值”，item_row+1行写入“同比yyyy-1”，item_row+2行写入“同比2019”
            for item_row in items_rows:
                ws.cell(row=item_row, column=items_col_num, value=f"当期值")
                ws.cell(row=item_row+1, column=items_col_num, value=f"同比{unified_date.year-1}")
                ws.cell(row=item_row+2, column=items_col_num, value=f"同比2019")
            print(f"✅ 参数列已写入")

        else:
            print("⚠️ 未找到任何有效的最新日期")
        
    @staticmethod
    def aviation_write_quarterly_report_header_info(context, params):
        """
        处理季度汇总表，写入标题与参数列。
        """
        wb = context['wb']
        ws = context['ws']
        
        # 获取目标工作表名称列表
        target_sheets_name = params['target_sheets']
        # 获取标题列与行数，参数列与参数行数
        title_col = params['title_col']
        title_col_num = column_letter_to_number(title_col)
        title_row = params['title_row']
        items_col = params['items_col']
        items_col_num = column_letter_to_number(items_col)
        items_rows = params.get('items_rows', [])
        
        # 列出目标文件中的所有工作表名称
        sheet_names = wb.sheetnames
        
        # 存储每个工作表的最新季度
        latest_quarters = {}
        latest_rows = {}

        for sheet_name in target_sheets_name:
            if sheet_name in sheet_names:
                ws_sheet = wb[sheet_name]
                
                # 遍历 A 列，获取最新季度标识
                quarters = []
                for row_idx in range(1, ws_sheet.max_row + 1):
                    cell_value = ws_sheet.cell(row=row_idx, column=1).value  # A 列
                    if cell_value and isinstance(cell_value, str) and 'Q' in cell_value:
                        # 检查是否符合 yyQn 格式
                        if cell_value.replace('Q', '').isdigit():
                            quarters.append(cell_value)
                
                if quarters:
                    # 获取最新的季度（假设按顺序排列，最后一个是最新的）
                    latest_quarter = quarters[-1]
                    # 获取最新季度的行数
                    latest_row = next((row_idx for row_idx in range(1, ws_sheet.max_row + 1) 
                                    if ws_sheet.cell(row=row_idx, column=1).value == latest_quarter), None)
                    latest_rows[sheet_name] = latest_row
                    latest_quarters[sheet_name] = latest_quarter
                    print(f"    - 工作表 [{sheet_name}] 的最新季度: {latest_quarter}")
                else:
                    print(f"    - 工作表 [{sheet_name}] 的 A 列没有有效季度数据")
            else:
                print(f"⚠️ 跳过不存在的工作表: {sheet_name}")
        
        # 检查所有工作表的最新季度是否一致
        if latest_quarters:
            ws = context['ws']
            unified_quarter = list(latest_quarters.values())[0]  # 取第一个工作表的最新季度作为基准
            if all(quarter == unified_quarter for quarter in latest_quarters.values()):
                ws.cell(row=1, column=column_letter_to_number('C'), value=list(latest_rows.values())[0])  # 将list(latest_rows.values())[0]写入C1
                print(f"✅ 所有工作表的最新季度一致: {unified_quarter}")
                
                # 解析季度字符串，例如 "25Q1" -> 年份2025，季度1
                try:
                    year_part = unified_quarter.split('Q')[0]
                    quarter_part = unified_quarter.split('Q')[1]
                    
                    # 假设是20xx年格式
                    full_year = f"20{year_part}"
                    
                    # 写入标题单元格
                    title = f"【申万交运】{full_year}年第{quarter_part}季度航空公司数据汇总"
                    ws.cell(row=title_row, column=title_col_num, value=title)
                    print(f"✅ 标题已写入: {title}")

                    # 写入参数列
                    # 对于 items_col_num 列，对于items_rows中的item_row，写入"当期值"，item_row+1行写入"同比yy-1Qn"，item_row+2行写入"同比2019Qn"
                    for item_row in items_rows:
                        ws.cell(row=item_row, column=items_col_num, value=f"当期值")
                        
                        # 计算去年同期季度
                        prev_year = int(year_part) - 1
                        yoy_quarter = f"{prev_year:02d}Q{quarter_part}"
                        ws.cell(row=item_row+1, column=items_col_num, value=f"同比{yoy_quarter}")
                        
                        # 同比2019年对应季度
                        ws.cell(row=item_row+2, column=items_col_num, value=f"同比19Q{quarter_part}")
                    print(f"✅ 参数列已写入")
                except Exception as e:
                    print(f"❌ 解析季度字符串失败: {e}")
                    raise ValueError(f"无法解析季度字符串: {unified_quarter}")
            else:
                print("❌ 工作表的最新季度不一致:")
                for sheet, quarter in latest_quarters.items():
                    print(f"    - {sheet}: {quarter}")
                raise ValueError("工作表的最新季度不一致，请检查数据！")
        else:
            print("⚠️ 未找到任何有效的最新季度")

    @staticmethod
    def aviation_write_report_data(context, params):
        """
        功能说明：
            在汇总表中写入跨表引用的公式。
            从总表第1行查找目标表名对应的列，然后根据目标行和源列生成 INDIRECT 跨表引用公式，
            实现从各子表自动汇总数据到总表的功能。

        params:
            context: dict — Pipeline 上下文，包含 ws。
            params: dict — 参数配置，包含：
                - target_rows: list[int] — 汇总表中待写入的行号列表。
                - target_sheets: list[str] — 目标子表名称列表。
                - target_sheet_columns: list[str] — 目标子表中需要读取的列字母列表。
        """        
        ws = context['ws']

        target_rows = params['target_rows']  # 待写入行
        target_sheets_name = params['target_sheets']  # 目标表
        target_sheet_columns = params['target_sheet_columns']  # 目标表需处理列

        # 目标表中需要读取的行数
        target_sheet_row = ws.cell(row=1, column=column_letter_to_number('C')).value

        for target_sheet_name in target_sheets_name:
            # 1. 在总表第 1 行中找到目标表名对应的列
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == target_sheet_name:
                    target_col = col
                    break

            if target_col is None:
                print(f"未在当前表第1行找到目标表名：{target_sheet_name}")
                continue

            # 2. 获取目标列字母
            target_col_letter = column_number_to_letter(target_col)

            # 3. 逐个写入公式
            for i, target_row in enumerate(target_rows):
                source_col_letter = target_sheet_columns[i].upper()

                # 生成公式
                formula = (
                    f'=@INDIRECT("\'"&{target_col_letter}$1&"\'!'
                    f'{source_col_letter}{target_sheet_row}")'
                )

                ws.cell(
                    row=target_row,
                    column=target_col
                ).value = formula
            # 4 清空C1
            ws.cell(row=1, column=column_letter_to_number('C')).value = None


