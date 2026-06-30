import re
import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num
from openpyxl.utils import get_column_letter


class ElectronicsPlugin:
    @staticmethod
    def _parse_indicators(indicators_config):
        """将 YAML 中的 indicators 配置统一转为 (code, col) 对列表。"""
        if isinstance(indicators_config, dict):
            return list(indicators_config.items())
        pairs = []
        for item in indicators_config:
            if isinstance(item, dict):
                pairs.append((item['code'], item['col']))
            elif isinstance(item, list):
                pairs.append((item[0], item[1]))
        return pairs

    @staticmethod
    def _parse_cell_date(value):
        """从单元格值解析为 datetime，无法解析返回 None。"""
        if value is None:
            return None
        if hasattr(value, 'strftime'):
            return value
        if isinstance(value, str) and value.strip():
            try:
                return pd.to_datetime(value)
            except Exception:
                pass
        return None

    @staticmethod
    def electronics_write_sheet(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        data_start_row = sheet_config.get('data_start_row', 51)

        sections = sheet_config.get('sections', [])

        for section in sections:
            date_col = section['date_col']
            date_col_num = col_letter_to_num(date_col)
            indicator_pairs = ElectronicsPlugin._parse_indicators(section['indicators'])

            # 按 AA code 去重加载数据
            all_codes = set(code for code, _ in indicator_pairs)
            code_dfs = {}
            for aa_code in all_codes:
                ind_data = reader.read_indicator_data(source_sheet, aa_code)
                df = ind_data["data"]
                if df.empty:
                    continue
                valid_dates = pd.to_datetime(df['日期'], errors='coerce')
                df = df[valid_dates.notna()].copy()
                value_col = df.columns[-1]
                df = df[df[value_col].notna()].copy()
                if df.empty:
                    continue
                df = df.sort_values('日期', ascending=True).reset_index(drop=True)
                code_dfs[aa_code] = df

            if not code_dfs:
                continue

            # 合并所有指标的日期并集（从新到旧排列）
            all_dates = set()
            for df in code_dfs.values():
                all_dates.update(df['日期'])
            date_list = sorted(all_dates, reverse=True)

            for i, date_val in enumerate(date_list):
                row = data_start_row + i
                date_cell = ws.cell(row=row, column=date_col_num)
                date_cell.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
                date_cell.number_format = 'yyyy-mm-dd'

            # 写入各指标值（对齐到统一日期列）
            for aa_code, target_col in indicator_pairs:
                df = code_dfs.get(aa_code)
                if df is None or df.empty:
                    continue
                value_map = dict(zip(df['日期'], df[df.columns[-1]]))
                target_col_num = col_letter_to_num(target_col)
                for i, date_val in enumerate(date_list):
                    row = data_start_row + i
                    val = value_map.get(date_val)
                    if val is not None and pd.notna(val):
                        cell = ws.cell(row=row, column=target_col_num)
                        cell.value = float(val)
                        cell.number_format = '0.00'

            print(f"    - 完成 {source_sheet} -> section [{date_col}] 共 {len(indicator_pairs)} 个指标")

    @staticmethod
    def electronics_update_chart_ranges(context, params):
        """更新图表引用范围：仅显示最近两年数据，同一图表内所有 series 统一起止行。"""
        ws = context['ws']
        sheet_config = context['sheet_config']
        data_start = sheet_config.get('data_start_row', 51)

        if not ws._charts:
            return

        for chart in ws._charts:
            # 解析所有 series 的列引用
            series_info = []
            for series in chart.series:
                if not (series.val and series.val.numRef and series.val.numRef.f):
                    continue
                m = re.match(
                    r"^(?:'?([^!']+)'?!)?\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)$",
                    series.val.numRef.f.strip(),
                )
                if m:
                    sheet_ref, col1, _, col2, _ = m.groups()
                    series_info.append((series, sheet_ref, col1, col2))

            if not series_info:
                continue

            # 日期列为第一个 series 值列左边一列
            date_col_num = max(col_letter_to_num(series_info[0][2]) - 1, 1)

            # 最新日期
            latest_date = ElectronicsPlugin._parse_cell_date(
                ws.cell(row=data_start, column=date_col_num).value
            )

            # 找到最后一个有数值的行
            cutoff_row = data_start
            for _, _, col1, _ in series_info:
                col_num = col_letter_to_num(col1)
                for r in range(ws.max_row, data_start - 1, -1):
                    v = ws.cell(row=r, column=col_num).value
                    if v is not None and str(v).strip() != '':
                        if r > cutoff_row:
                            cutoff_row = r
                        break

            # 应用两年截断：向下找到第一个 <= 两年前的日期行
            if latest_date is not None:
                cutoff_date = latest_date - pd.DateOffset(years=2)
                for r in range(data_start, cutoff_row + 1):
                    dt = ElectronicsPlugin._parse_cell_date(
                        ws.cell(row=r, column=date_col_num).value
                    )
                    if dt is not None and dt <= cutoff_date:
                        cutoff_row = r
                        break

            date_col_letter = get_column_letter(date_col_num)

            # 统一重写所有 series 的值和类别范围
            for series, sheet_ref, col1, col2 in series_info:
                prefix = f"'{sheet_ref}'!" if sheet_ref else ""
                new_val_f = f"{prefix}${col1}${data_start}:${col2}${cutoff_row}"
                new_cat_f = f"{prefix}${date_col_letter}${data_start}:${date_col_letter}${cutoff_row}"

                series.val.numRef.f = new_val_f
                if hasattr(series.val.numRef, 'numCache'):
                    series.val.numRef.numCache = None
                if series.cat and series.cat.numRef and series.cat.numRef.f:
                    series.cat.numRef.f = new_cat_f
                    if hasattr(series.cat.numRef, 'numCache'):
                        series.cat.numRef.numCache = None

        print(f"    - 完成图表数据范围更新（最近两年）")
