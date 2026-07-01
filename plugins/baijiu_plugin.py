import re
import zipfile
import shutil
import pandas as pd
from openpyxl.utils import column_index_from_string as col_letter_to_num


class BaijiuPlugin:
    """白酒 Pipeline 插件：按产品名匹配写入 + 图表 XML 更新"""

    # 数据写入：按指标代码匹配（源指标代码→目标列）
    # ================================================================
    @staticmethod
    def baijiu_write_sheet_by_code(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sc = context['sheet_config']

        source_sheet = sc['source_sheet']
        data_start = sc.get('data_start_row', 9)
        date_col_num = col_letter_to_num(sc['date_col'])
        indicator_pairs = list(sc.get('indicators', {}).items())

        all_codes = set(code for code, _ in indicator_pairs)
        code_dfs = {}
        for aa_code in all_codes:
            ind_data = reader.read_indicator_data(source_sheet, aa_code)
            df = ind_data["data"]
            if df.empty:
                continue
            valid_dates = pd.to_datetime(df['日期'], errors='coerce')
            df = df[valid_dates.notna()].copy()
            value_col = df.columns[-1]
            df = df[df[value_col].notna()].copy()
            if df.empty:
                continue
            df = df.sort_values('日期', ascending=True).reset_index(drop=True)
            code_dfs[aa_code] = df

        if not code_dfs:
            print(f"    - [{source_sheet}] 无有效指标数据")
            return

        # 合并所有指标的日期并集
        all_dates = set()
        for df in code_dfs.values():
            all_dates.update(df['日期'])
        date_list = sorted(all_dates, reverse=False)

        # 写日期列
        for i, date_val in enumerate(date_list):
            row = data_start + i
            date_cell = ws.cell(row=row, column=date_col_num)
            date_cell.value = date_val.to_pydatetime() if isinstance(date_val, pd.Timestamp) else date_val
            date_cell.number_format = 'yyyy-mm-dd'

        # 写各指标值
        for aa_code, target_col in indicator_pairs:
            df = code_dfs.get(aa_code)
            if df is None or df.empty:
                continue
            value_map = dict(zip(df['日期'], df[df.columns[-1]]))
            target_col_num = col_letter_to_num(target_col)
            for i, date_val in enumerate(date_list):
                row = data_start + i
                val = value_map.get(date_val)
                if val is not None and pd.notna(val):
                    cell = ws.cell(row=row, column=target_col_num)
                    try:
                        cell.value = float(val)
                    except (ValueError, TypeError):
                        cell.value = val
                    cell.number_format = '0.00'

        print(f"    - [{source_sheet}] {len(indicator_pairs)}个指标代码匹配写入")

    # ================================================================
    # 汇总行填充：计算最新/上周/上月/年初值，写入数据 sheet 顶部行
    # 取代原 VLOOKUP 公式（openpyxl 不会计算，保存后均为 None）
    # ================================================================
    @staticmethod
    def baijiu_fill_summary_rows(context, params):
        ws = context['ws']
        sc = context['sheet_config']
        from datetime import timedelta

        date_col_num = col_letter_to_num(sc['date_col'])
        data_start = sc.get('data_start_row', 7)

        # 收集所有数据列 (有非空值的列)
        data_cols = []
        for c in range(date_col_num + 1, ws.max_column + 1):
            for r in range(data_start, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    data_cols.append(c)
                    break

        if not data_cols:
            return

        # 建立日期→行号映射
        date_rows = {}
        all_dates = []
        for r in range(data_start, ws.max_row + 1):
            dv = ws.cell(row=r, column=date_col_num).value
            if dv is not None:
                dt = pd.to_datetime(dv) if not hasattr(dv, 'strftime') else dv
                date_rows[r] = dt
                all_dates.append(dt)

        if not all_dates:
            return

        latest = max(all_dates)
        targets = {
            1: latest,
            2: latest - timedelta(days=7),
            3: latest - timedelta(days=28),
            4: latest - timedelta(days=364),
        }

        for col in data_cols:
            col_values = {}
            for r in range(data_start, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    col_values[r] = v

            for tgt_row, tgt_date in targets.items():
                best_row = None
                best_date = None
                for r, d in date_rows.items():
                    if d <= tgt_date and r in col_values:
                        if best_date is None or d > best_date:
                            best_date = d
                            best_row = r
                if best_row is not None:
                    cell = ws.cell(row=tgt_row, column=col)
                    cell.value = col_values[best_row]
                    cell.number_format = '0.00'

        # 填充日期列（用于 G1, G2, G3, G4）
        date_output_col = date_col_num - 1  # 通常在日期列左边一列
        for tgt_row, tgt_date in targets.items():
            cell = ws.cell(row=tgt_row, column=date_output_col)
            cell.value = tgt_date.to_pydatetime() if hasattr(tgt_date, 'to_pydatetime') else tgt_date
            cell.number_format = 'yyyy-mm-dd'

        print(f"    - [{sc['sheet_name']}] 汇总行填充完成")

    @staticmethod
    def baijiu_write_sheet(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sc = context['sheet_config']

        source_sheet = sc['source_sheet']
        data_start = sc.get('data_start_row', 9)
        date_col_num = col_letter_to_num(sc['date_col'])

        df = reader.read_sheet_data(source_sheet)

        # --- 1. 读取源产品名 → source_col_idx ---
        meta = df.iloc[:10]
        src_products = {}  # product_name → 0-based df col index
        src_brands = {}
        for c in range(1, len(df.columns)):
            full = str(meta.iloc[0, c]) if pd.notna(meta.iloc[0, c]) else ''
            parts = full.split(':')
            brand = parts[3] if len(parts) > 3 else ''
            pname = parts[4] if len(parts) > 4 else ''
            if pname:
                src_products[pname] = c
                src_brands[pname] = brand

        # --- 2. 读取目标表头 → target_col_num ---
        tgt_headers = {}  # header_text → target_col_num
        for c in range(date_col_num + 1, ws.max_column + 1):
            hdr = ws.cell(row=sc.get('header_read_row', data_start - 2), column=c).value
            if hdr:
                tgt_headers[str(hdr).strip()] = c
            # Also check row above (brand row)
            hdr2 = ws.cell(row=sc.get('header_read_row', data_start - 2) - 1, column=c).value
            if hdr2:
                tgt_headers[str(hdr2).strip()] = c

        # --- 3. 手动映射表 ---
        # Config 格式：{源产品名: 目标表头}
        raw_map = sc.get('manual_map', {})
        manual_col_map = {}  # tgt_col_num → src_col_idx
        for src_pname, tgt_hdr in raw_map.items():
            # 查找源列
            src_idx = src_products.get(src_pname)
            if src_idx is None:
                continue
            # 查找目标列
            tgt_col = tgt_headers.get(tgt_hdr)
            if tgt_col is None:
                continue
            manual_col_map[tgt_col] = src_idx

        # --- 4. 匹配：目标表头 → 源列索引 ---
        col_map = dict(manual_col_map)  # 先手动映射
        used_src = set(col_map.values())

        # 跳过日期/品牌行
        data_headers = [(k, v) for k, v in tgt_headers.items()
                        if not any(kw in k for kw in ['日期', '从早', '从新'])]

        # 第一遍：精确名称匹配
        rest = []
        for tgt_name, tgt_col in data_headers:
            if tgt_col in col_map:
                continue
            if tgt_name in src_products and src_products[tgt_name] not in used_src:
                col_map[tgt_col] = src_products[tgt_name]
                used_src.add(src_products[tgt_name])
            else:
                rest.append((tgt_name, tgt_col))

        # 第二遍：模糊匹配
        for tgt_name, tgt_col in rest:
            if tgt_col in col_map:
                continue
            for sp, sc_idx in src_products.items():
                if sc_idx in used_src:
                    continue
                if sp in tgt_name or tgt_name in sp:
                    col_map[tgt_col] = sc_idx
                    used_src.add(sc_idx)
                    break

        print(f"    - [{source_sheet}] {len(col_map)}列匹配")

        # --- 5. 数据处理 ---
        data_df = df.iloc[10:].reset_index(drop=True)
        if data_df.empty:
            return

        data_df.columns = [f'col_{i}' for i in range(len(data_df.columns))]
        data_df = data_df.rename(columns={'col_0': 'date'})
        data_df['date'] = pd.to_datetime(data_df['date'], errors='coerce')
        data_df = data_df[data_df['date'].notna()].copy()
        data_df = data_df.sort_values('date', ascending=True).reset_index(drop=True)

        # --- 6. 写日期 ---
        for i, dv in enumerate(data_df['date']):
            c = ws.cell(row=data_start + i, column=date_col_num)
            c.value = dv.to_pydatetime() if isinstance(dv, pd.Timestamp) else dv
            c.number_format = 'yyyy-mm-dd'

        # --- 7. 写数据 ---
        for tgt_col, src_col_idx in col_map.items():
            src_col_name = f'col_{src_col_idx}'
            if src_col_name not in data_df.columns:
                continue
            for i, val in enumerate(data_df[src_col_name]):
                if pd.notna(val):
                    c = ws.cell(row=data_start + i, column=tgt_col)
                    try:
                        c.value = float(val)
                    except (ValueError, TypeError):
                        c.value = val
                    c.number_format = '0.00'

    # ================================================================
    # 图表 XML 更新
    # ================================================================
    _CHART_SHEET_MAP = {
        4: -1, 6: -1, 8: -1,   # 作图 sheet 自身（用虚拟配置 N=date_col）
        5: 2, 14: 2,  # 飞天次新酒
        7: 3,   # 白酒品牌批价（今日酒价）
        9: 1,   # 飞天100ml 和 飞天1L原箱
        10: 0, 11: 0, 12: 0,   # 各酒厂主要产品
        13: 1,  # 飞天100ml 和 飞天1L原箱
        15: 3, 16: 3, 17: 3, 18: 3, 19: 3, 20: 3, 21: 3, 22: 3, 23: 3,  # 白酒品牌批价（今日酒价）
        24: 3, 25: 3, 26: 3, 27: 3, 28: 3, 29: 3, 30: 3, 31: 3,
        32: 4, 33: 4, 34: 4, 35: 4, 36: 4, 37: 4, 38: 4, 39: 4, 40: 4,  # 白酒品牌批价 (酒价参考)
        41: 4, 42: 4, 43: 4, 44: 4, 45: 4, 46: 4, 47: 4, 48: 4,
    }

    @staticmethod
    def baijiu_finalize_charts(context, params):
        filepath = context['filepath']
        config = context['config']
        tmp = filepath + '.tmp'

        # 从原始模板恢复被 openpyxl 破坏的 drawing/vml 文件
        template_file = config.get('target_file', filepath)
        with zipfile.ZipFile(template_file, 'r') as ztpl:
            tpl_data = {name: ztpl.read(name) for name in ztpl.namelist()}
        tpl_names = set(tpl_data.keys())

        with zipfile.ZipFile(filepath, 'r') as zin, \
             zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            out_names = set(zin.namelist())

            for item in zin.infolist():
                fn = item.filename
                data = zin.read(fn)

                # 恢复被 openpyxl 破坏的 drawing/vml/theme (含 _rels)
                if (fn.startswith('xl/drawings/') or
                    fn.startswith('xl/ctrlProps/') or
                    fn.startswith('xl/theme/')) and fn in tpl_names:
                    data = tpl_data[fn]

                # 修补图表 XML
                if fn.startswith('xl/charts/chart') and fn.endswith('.xml'):
                    data = BaijiuPlugin._patch_chart_xml(data, fn, filepath, config)

                # openpyxl 把 vmlDrawing 改名成了 commentsDrawing，跳过它
                if fn.startswith('xl/drawings/commentsDrawing') and fn.endswith('.vml'):
                    continue

                zout.writestr(item, data)

            # 补充 openpyxl 删除的文件
            for tpl_name in tpl_names:
                if (tpl_name.startswith('xl/drawings/vmlDrawing') or
                    tpl_name.startswith('xl/ctrlProps/') or
                    tpl_name.startswith('xl/theme/')) and tpl_name not in out_names:
                    info = zipfile.ZipInfo(tpl_name)
                    info.compress_type = zipfile.ZIP_DEFLATED
                    zout.writestr(info, tpl_data[tpl_name])

        shutil.move(tmp, filepath)
        print(f"    [XML] 图表更新完成（含模板图像恢复）")

    @staticmethod
    def _patch_chart_xml(xml_bytes, filename, filepath, config):
        xml = xml_bytes.decode('utf-8')
        m = re.search(r'chart(\d+)\.xml', filename)
        if not m:
            return xml_bytes
        cn = int(m.group(1))

        sheet_idx = BaijiuPlugin._CHART_SHEET_MAP.get(cn)
        if sheet_idx is None:
            return xml_bytes
        if sheet_idx == -1:
            sc = {'sheet_name': '作图', 'date_col': 'N', 'data_start_row': 7, 'max_data_cols': 16}
            is_first_two = False
        else:
            sc = config['sheets'][sheet_idx]
            is_first_two = (sheet_idx in [0, 1])

        cs, er = BaijiuPlugin._chart_rows(filepath, sc, is_first_two)
        if cs is None:
            return xml_bytes

        if sc.get('max_data_cols', 0):
            max_tc = col_letter_to_num(sc['date_col']) + sc['max_data_cols']
            xml = BaijiuPlugin._strip_oob(xml, max_tc)

        xml = re.sub(
            r'\x24([A-Z]+)\x24(\d+):\x24([A-Z]+)\x24(\d+)',
            lambda m: f'${m.group(1)}${cs}:${m.group(3)}${er}',
            xml
        )

        if not is_first_two:
            xml = BaijiuPlugin._set_axis(xml)

        for tag in ['dLbls', 'numCache', 'strCache']:
            xml = re.sub(f'<(?:c:)?{tag}>.*?</(?:c:)?{tag}>', '', xml, flags=re.DOTALL)
            xml = re.sub(f'<(?:c:)?{tag}[^/>]*/>', '', xml)

        if cn == 10:
            xml = re.sub(r'<(?:c:)?min val="[^"]*"/>', '', xml)
            xml = re.sub(r'<(?:c:)?max val="[^"]*"/>', '', xml)

        return xml.encode('utf-8')

    @staticmethod
    def _strip_oob(xml, max_tc):
        pat = r'\x24([A-Z]+)\x24(\d+):\x24([A-Z]+)\x24(\d+)'
        def replacer(m):
            all_m = re.findall(pat, m.group(0))
            if all_m:
                try:
                    if col_letter_to_num(all_m[-1][0]) > max_tc:
                        return ''
                except:
                    pass
            return m.group(0)
        return re.sub(r'<ser>.*?</ser>', replacer, xml, flags=re.DOTALL)

    @staticmethod
    def _set_axis(xml):
        if '<majorUnit' in xml:
            xml = re.sub(r'<majorUnit val="[^"]*"/>', '<majorUnit val="365"/>', xml)
        else:
            xml = re.sub(r'<dateAx>', r'<dateAx><majorUnit val="365"/><minorUnit val="30"/>', xml)
        if '<minorUnit' not in xml:
            xml = re.sub(r'(<majorUnit[^/]+/>)', r'\1<minorUnit val="30"/>', xml)
        return xml

    @staticmethod
    def _chart_rows(filepath, sc, is_first_two):
        data_start = sc.get('data_start_row', 9)
        dc = col_letter_to_num(sc['date_col'])
        sheet_name = sc.get('sheet_name', '')

        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None, None
        ws = wb[sheet_name]

        end_row, latest = data_start, None
        for r in range(ws.max_row, data_start - 1, -1):
            v = ws.cell(row=r, column=dc).value
            if v is not None and str(v).strip() != '':
                if end_row == data_start:
                    end_row = r
                if latest is None:
                    try:
                        latest = pd.to_datetime(v)
                    except Exception:
                        pass
                if end_row != data_start and latest is not None:
                    break

        chart_start = data_start
        if latest is not None and is_first_two:
            cutoff = pd.Timestamp(year=latest.year - 2, month=1, day=1)
            for r in range(data_start, end_row + 1):
                v = ws.cell(row=r, column=dc).value
                if v is not None:
                    try:
                        if pd.to_datetime(v) >= cutoff:
                            chart_start = r
                            break
                    except Exception:
                        pass

        wb.close()
        return chart_start, end_row

    # ================================================================
    # 作图 Sheet 填充：解析公式引用，写入实算值
    # ================================================================
    @staticmethod

    # ================================================================
    @staticmethod
    def baijiu_fill_zuotu(context, params):
        filepath = context['filepath']
        import openpyxl
        wb = openpyxl.load_workbook(filepath)

        sheet_name = None
        for sn in wb.sheetnames:
            if len(sn) == 2 and sn.encode('utf-8') == b'\xe4\xbd\x9c\xe5\x9b\xbe':
                sheet_name = sn
                break
        if sheet_name is None:
            wb.close()
            return

        ws = wb[sheet_name]

        def resolve_any_cell(rws, cell_addr):
            clean = cell_addr.replace('$', '')
            val = rws[clean].value
            if val is None:
                return None
            if not isinstance(val, str) or not val.startswith('='):
                return val
            return _resolve_formula(val, rws)

        def _resolve_formula(f, current_ws):
            f = f.strip()
            if not f.startswith('='):
                return f
            # Strip [NN] external workbook references
            f_noeq = f[1:]
            f_noeq = re.sub(r'\[\d+\]', '', f_noeq)
            # VLOOKUP
            vm = re.match(
                                r"^VLOOKUP\(\s*(.+?)\s*,\s*\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)\s*,\s*(\d+)\s*,\s*(TRUE|FALSE)\s*\)$",
                f_noeq, re.IGNORECASE
            )
            if vm:
                lookup_expr, sc_col, sc_row, ec_col, ec_row, col_idx_str, approx = vm.groups()
                lookup_val = _resolve_expr(lookup_expr, current_ws)
                if lookup_val is None:
                    return None
                return _vlookup(lookup_val, vm.group(2) + vm.group(3), vm.group(4) + vm.group(5),
                                int(col_idx_str), approx.upper() == 'TRUE', current_ws)
            # Cross-sheet ref
            ref_m = re.match(r"^(?:'([^']+)'|([A-Za-z0-9_一-鿿 ()（）\-]+))!(\$?[A-Z]+\$?\d+)$", f_noeq)
            if ref_m:
                ref_sheet = ref_m.group(1) or ref_m.group(2)
                ref_addr = ref_m.group(3).replace('$', '')
                if ref_sheet in wb.sheetnames:
                    return resolve_any_cell(wb[ref_sheet], ref_addr)
            # MAX
            max_m = re.match(r"^MAX\(([A-Z]+\d+):([A-Z]+\d+)\)$", f_noeq, re.IGNORECASE)
            if max_m:
                return _range_max(max_m.group(1), max_m.group(2), current_ws)
            # Arithmetic: =EXPR1-EXPR2 with optional sheet refs
            sheet_pat = r"(?:(?:'[^']*')|[^!\-+*/^&()=<>,]+)!"
            cell_only = r"\$?[A-Z]+\$?\d+"
            cell_with_sheet = rf"(?:{sheet_pat})?{cell_only}"
            am = re.match(rf"^({cell_with_sheet})-({cell_with_sheet})$", f_noeq)
            if am:
                c1 = _resolve_expr(am.group(1), current_ws)
                c2 = _resolve_expr(am.group(2), current_ws)
                if c1 is not None and c2 is not None:
                    try:
                        return float(c1) - float(c2)
                    except (ValueError, TypeError):
                        return None
            return None

        def _resolve_expr(expr, current_ws):
            expr = expr.strip()
            if expr.startswith("="):
                return _resolve_formula(expr, current_ws)
            # Match cell ref from the right: always ends with COL$ROW
            cell_m = re.search(r"(\$?[A-Z]+\$?\d+)$", expr)
            if cell_m:
                cell_part = cell_m.group(1).replace("$", "")
                sheet_part = expr[:cell_m.start()]
                if sheet_part:
                    if sheet_part.endswith("!"):
                        sheet_part = sheet_part[:-1]
                    sheet_part = sheet_part.strip("'").strip('"')
                    if sheet_part in wb.sheetnames:
                        return resolve_any_cell(wb[sheet_part], cell_part)
                else:
                    return resolve_any_cell(current_ws, cell_part)
            try:
                return float(expr)
            except (ValueError, TypeError):
                pass
            try:
                return pd.to_datetime(expr)
            except Exception:
                pass
            return expr

        def _vlookup(lookup_val, start_cell, end_cell, col_idx, is_approx, current_ws):
            from openpyxl.utils import column_index_from_string as _cn
            sc_col = re.match(r"^([A-Z]+)", start_cell).group(1)
            sc_row = int(re.match(r"^[A-Z]+(\d+)", start_cell).group(1))
            ec_col = re.match(r"^([A-Z]+)", end_cell).group(1)
            ec_row = int(re.match(r"^[A-Z]+(\d+)", end_cell).group(1))
            lookup_col = _cn(sc_col)
            target_col = lookup_col + col_idx - 1
            if target_col > _cn(ec_col):
                return None
            best_row, best_val = None, None
            import pandas as pd
            for r in range(sc_row, ec_row + 1):
                lv = current_ws.cell(row=r, column=lookup_col).value
                if lv is None:
                    continue
                if isinstance(lv, str):
                    try:
                        lv = float(lv)
                    except ValueError:
                        try:
                            lv = pd.to_datetime(lv)
                        except Exception:
                            pass
                if isinstance(lookup_val, pd.Timestamp) and not isinstance(lv, pd.Timestamp):
                    try:
                        lv = pd.to_datetime(lv)
                    except Exception:
                        pass
                elif isinstance(lookup_val, (int, float)) and not isinstance(lv, (int, float)):
                    try:
                        lv = float(lv)
                    except (ValueError, TypeError):
                        pass
                if not is_approx:
                    if lv == lookup_val:
                        v = current_ws.cell(row=r, column=target_col).value
                        return float(v) if v is not None else None
                else:
                    if lv <= lookup_val:
                        if best_val is None or lv > best_val:
                            best_val = lv
                            best_row = r
            if best_row is not None:
                v = current_ws.cell(row=best_row, column=target_col).value
                if v is not None:
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return v
            return None

        def _range_max(start_cell, end_cell, current_ws):
            from openpyxl.utils import column_index_from_string as _cn
            sc_col = re.match(r"^([A-Z]+)", start_cell).group(1)
            sc_row = int(re.match(r"^[A-Z]+(\d+)", start_cell).group(1))
            ec_col = re.match(r"^([A-Z]+)", end_cell).group(1)
            ec_row = int(re.match(r"^[A-Z]+(\d+)", end_cell).group(1))
            max_val = None
            for r in range(sc_row, ec_row + 1):
                for c in range(_cn(sc_col), _cn(ec_col) + 1):
                    v = current_ws.cell(row=r, column=c).value
                    if v is not None:
                        try:
                            fv = float(v)
                            if max_val is None or fv > max_val:
                                max_val = fv
                        except (ValueError, TypeError):
                            pass
            return max_val

        max_c = params.get('max_column', 13)
        max_r = params.get('max_row', 50)
        count = 0
        import pandas as pd
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    val = _resolve_formula(cell.value, ws)
                    if val is not None:
                        cell.value = val
                        if isinstance(val, (int, float)):
                            cell.number_format = '0.00'
                        elif isinstance(val, pd.Timestamp):
                            cell.number_format = 'yyyy-mm-dd'
                        count += 1

        wb.save(filepath)
        wb.close()
        import pandas as pd
        print(f"    [zuotu] filled {count} formula values")
