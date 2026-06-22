import pandas as pd
import openpyxl
import re
from openpyxl.utils import column_index_from_string as column_letter_to_number, get_column_letter as column_number_to_letter
from datetime import datetime

# 为了重用之前的复杂代码，直接引入旧的 DataProcessor (作为示例平滑迁移，实际可全盘重构进这里)
from core_engine.data_processor import DataProcessor

class BankPlugin:
    @staticmethod
    def bank_write_data(context, params):
        print(f"开始处理银行数据")
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 50)
        start_column = params.get('start_column', 2)
        start_date = params.get('start_date', '2021-01-01')
        date_format = params.get('date_format', 'yyyy-mm')
        start_date_ts = pd.to_datetime(start_date)

        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期', ascending=False).reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表，并筛选季度数据（3、6、9、12月）
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        # 筛选只包含3、6、9、12月的数据
        quarterly_df = first_df[first_df['日期'].dt.month.isin([3, 6, 9, 12])].copy()
        
        if quarterly_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 中没有3、6、9、12月的数据，跳过写入")
            return

        base_df = quarterly_df.reset_index(drop=True)

        # 3. 写入统一日期列和指标数据（应用单位转换）
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            # 第1列写日期
            date_cell = ws.cell(row=current_row, column=start_column, value=current_date)
            date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            # 对每个指标也筛选季度数据
            df_quarterly = df[df['日期'].dt.month.isin([3, 6, 9, 12])].copy()
            
            if df_quarterly.empty:
                print(f"    - 指标 {indicator} 中没有3、6、9、12月的数据")
                continue
                
            value_col = df_quarterly.columns[-1]
            value_map = dict(zip(df_quarterly['日期'], df_quarterly[value_col]))
            col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)    
            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                target_cell = ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=value
                )
                
                # 继承同列首行的原始格式
                original_cell = ws.cell(row=start_row, column=col_num)
                target_cell.number_format = original_cell.number_format

        print(f"    - 完成银行基础数据写入")


    @staticmethod
    def bank_commercial_write_data(context, params):
        ws = context['ws']
        reader = context['data_reader']
        sheet_config = context['sheet_config']
        source_sheet = sheet_config['source_sheet']
        indicator_col_map = sheet_config['indicators']
        start_row = params.get('start_row', 50)
        start_column = params.get('start_column', 1)
        unit_conversion_row = params.get('unit_conversion', 0)
        start_date = params.get('start_date', '2014-01-01')
        date_format = params.get('date_format', 'yyyy-mm')
        # 新增：支持多列日期写入，可以是单个列号或列号列表
        date_columns = params.get('date_columns', None)
        start_date_ts = pd.to_datetime(start_date)

        def convert_value(value, unit_conversion):
            """根据列和配置的倍率转换数值"""
            if value is None or not isinstance(value, (int, float)):
                return value
            if unit_conversion:
                return value * unit_conversion
            return value
                
        # 1. 从缓存中获取所需所有指标数据
        indicator_dfs = {}
        for indicator in indicator_col_map:
            ind_data = reader.read_indicator_data(source_sheet, indicator)
            df = ind_data["data"].copy()

            if df.empty or '日期' not in df.columns:
                indicator_dfs[indicator] = pd.DataFrame(columns=['日期'])
                print(f"    - 工作表[{source_sheet}] 未找到指标 {indicator}，该列跳过写入")
                continue

            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')

            # 只保留 start_date 及以后数据
            df = df[df['日期'].notna() & (df['日期'] >= start_date_ts)].copy()
            df = df.sort_values('日期').reset_index(drop=True)

            indicator_dfs[indicator] = df

        # 2. 使用第一个指标的数据作为基础日期表（不再补全日期）
        first_indicator = list(indicator_col_map.keys())[0]
        first_df = indicator_dfs[first_indicator].copy()

        if first_df.empty:
            print(f"    - {source_sheet} 的第一个指标 {first_indicator} 无有效数据，跳过写入")
            return

        # 直接使用第一个指标的数据作为基础日期表
        base_df = first_df.copy()

        # 3. 写入统一日期列和指标数据（应用单位转换）
        # 确定需要写入日期的列列表
        if date_columns is not None:
            # 如果配置了 date_columns，转换为列表格式
            if isinstance(date_columns, int):
                date_cols_list = [date_columns]
            elif isinstance(date_columns, str):
                # 如果是字符串形式的列字母，转换为数字
                try:
                    date_cols_list = [column_letter_to_number(date_columns)]
                except:
                    date_cols_list = [int(date_columns)]
            else:
                # 已经是列表，处理每个元素
                date_cols_list = []
                for col in date_columns:
                    if isinstance(col, str):
                        try:
                            date_cols_list.append(column_letter_to_number(col))
                        except:
                            date_cols_list.append(int(col))
                    else:
                        date_cols_list.append(int(col))
        else:
            # 如果没有配置 date_columns，使用默认的 start_column
            date_cols_list = [start_column]
        
        for i, row in base_df.iterrows():
            current_row = start_row + i
            current_date = row['日期']

            # 在所有配置的日期列中写入日期
            for col_num in date_cols_list:
                if col_num >= 1:
                    date_cell = ws.cell(row=current_row, column=col_num, value=current_date)
                    date_cell.number_format = date_format

        # 4. 写入指标数据，全部按 base_df 的日期对齐（应用单位转换）
        base_date_list = base_df['日期'].tolist()

        for indicator, col in list(indicator_col_map.items()):
            df = indicator_dfs[indicator].copy()
            value_col = df.columns[-1]
            value_map = dict(zip(df['日期'], df[value_col]))
            
            # 添加安全检查
            try:
                col_num = column_letter_to_number(col) if isinstance(col, str) else int(col)
                if col_num < 1:
                    print(f"    - 跳过列 {col}，计算出的列号 {col_num} 无效")
                    continue
                    
                if unit_conversion_row >= 1 and col_num >= 1:
                    unit_cell = ws.cell(row=unit_conversion_row, column=col_num)
                    if unit_cell.value is not None and unit_cell.value > 0:
                        unit_conversion = unit_cell.value
                    else:
                        unit_conversion = 1
                else:
                    unit_conversion = 1
            except Exception as e:
                print(f"    - 获取单位转换因子失败，使用默认值1: {e}")
                unit_conversion = 1
                
            for i, date in enumerate(base_date_list):
                value = value_map.get(date, None)
                converted_value = convert_value(value, unit_conversion)
                target_cell = ws.cell(
                    row=start_row + i,
                    column=col_num,
                    value=converted_value
                )
                
                # 继承同列首行的原始格式
                original_cell = ws.cell(row=start_row, column=col_num)
                target_cell.number_format = original_cell.number_format
                
        return

    @staticmethod
    def bank_commercial_formula(context, params):
        """
        通用公式写入函数，支持自定义公式模板
        
        参数说明：
        - start_row: 起始行号
        - target_formula: 公式模板字符串，支持 {source_column}, {row}, {tname} 等占位符
        - target_column: 目标列列表（写入公式的列）
        - source_column: 源列列表（公式中引用的列）
        - tname: 可选的工作表名称，用于跨表引用
        - format: 可选的数字格式字符串，如 "0.00%"、"#,##0.00"、"[Red](#,##0.00)" 等
        
        """
        ws = context['ws']
        
        start_row = params.get('start_row', 6)
        target_formula_template = params.get('target_formula', '')
        target_columns = params.get('target_column', [])
        source_columns = params.get('source_column', [])
        tname = params.get('tname', None)  # 可选的工作表名称
        custom_format = params.get('format', None)  # 可选的自定义格式
        
        # 检查参数有效性
        if not target_formula_template:
            print("    - 缺少 target_formula 配置，跳过公式写入")
            return
            
        if not target_columns or not source_columns:
            print("    - 缺少 target_column 或 source_column 配置，跳过公式写入")
            return
            
        if len(target_columns) != len(source_columns):
            print(f"    - target_column 和 source_column 数量不匹配，跳过公式写入")
            return
        
        # 遍历每一对源列和目标列
        for target_col, source_col in zip(target_columns, source_columns):
            try:
                # 转换列标识为数字
                target_col_num = column_letter_to_number(target_col) if isinstance(target_col, str) else int(target_col)
                source_col_num = column_letter_to_number(source_col) if isinstance(source_col, str) else int(source_col)
                
                if target_col_num < 1 or source_col_num < 1:
                    print(f"    - 跳过无效的列配置: target={target_col}, source={source_col}")
                    continue
                
                # 从start_row开始，逐行写入公式，直到target_col的前一列（时间列）没有值为止
                current_row = start_row
                while True:
                    # 检查target_column-1的列是否有值（用于判断是否停止）
                    check_col_num = target_col_num - 1
                    if check_col_num < 1:
                        # 如果target是第1列，无法检查前一列，使用其他判断方式
                        check_cell = ws.cell(row=current_row, column=target_col_num)
                        if check_cell.value is None:
                            break
                    else:
                        check_cell = ws.cell(row=current_row, column=check_col_num)
                        if check_cell.value is None:
                            break
                    
                    # 构建公式：替换模板中的占位符
                    formula = target_formula_template
                    
                    # 替换 {source_column} 为实际的源列字母
                    source_col_letter = column_number_to_letter(source_col_num)
                    formula = formula.replace('{source_column}', source_col_letter)
                    
                    # 先替换所有 {row-N} 偏移占位符（需在 {row} 之前，避免误匹配）
                    formula = re.sub(r'\{row-(\d+)\}', lambda m: str(current_row - int(m.group(1))), formula)

                    # 替换 {row} 为当前行号
                    formula = formula.replace('{row}', str(current_row))
                    
                    # 如果有 tname，替换 {tname}
                    if tname:
                        formula = formula.replace('{tname}', tname)
                    
                    # 在目标单元格写入公式
                    target_cell = ws.cell(row=current_row, column=target_col_num)
                    target_cell.value = formula

                    # 设置单元格格式
                    if custom_format:
                        # 如果提供了自定义格式，直接使用
                        target_cell.number_format = custom_format
                    else:
                        # 否则从源列继承格式
                        source_cell = ws.cell(row=current_row, column=source_col_num)
                        target_cell.number_format = source_cell.number_format
                        
                        # 如果源单元格没有特定格式，则使用会计格式（负数用括号表示）
                        if not source_cell.number_format or source_cell.number_format == 'General':
                            target_cell.number_format = '#,##0.00_);(#,##0.00)'

                    current_row += 1
                
                print(f"    - 完成列 {target_col} 的公式写入（共 {current_row - start_row} 行）")
                
            except Exception as e:
                print(f"    - 处理列 {target_col} 时出错: {e}")
                continue
        
        return


